"""
IGREJA ABA — Backend v5
Melhorias: Permissões granulares, NAREAL, Edição de cultos,
           Busca de endereço fuzzy, PDF real, GC com líder,
           Estoque corrigido, Dashboard analytics, Logs
"""
from flask import Flask, request, jsonify, render_template, session, redirect, send_file, make_response
from flask_cors import CORS
import os, hashlib, secrets, io, qrcode, base64, urllib.parse, logging, math, json, re, sqlite3
try:
    import psycopg2, psycopg2.extras
    HAS_PG = True
except ImportError:
    HAS_PG = False
from datetime import datetime, date, timedelta
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder="static", static_url_path="/static",
            template_folder="templates")

_secret = os.environ.get("SECRET_KEY", "igreja-aba-key-2025-seguro")
app.secret_key = _secret
app.config.update(
    SESSION_COOKIE_SECURE   = bool(os.environ.get("RENDER","")),
    SESSION_COOKIE_HTTPONLY = True,
    SESSION_COOKIE_SAMESITE = "Lax",
    SEND_FILE_MAX_AGE_DEFAULT = 0,
    MAX_CONTENT_LENGTH = 16*1024*1024,
)
CORS(app, supports_credentials=True, origins="*")

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = os.environ.get("DATABASE_URL","").strip()
USE_PG       = bool(DATABASE_URL and HAS_PG)
_db_dir      = os.environ.get("DB_DIR","").strip() or os.path.join(BASE_DIR,"database")
DB_DIR       = _db_dir
DB_PATH      = os.path.join(DB_DIR,"igreja_aba.db")
SQL_PATH     = os.path.join(BASE_DIR,"database","schema.sql")

TIPOS_CULTO = ["Culto Regular","NAREAL","Evento","Reunião de Líderes","Culto de GC","Outro"]

# ── DB ────────────────────────────────────────────────────────

class _PGCursorWrapper:
    """Wrapper de cursor PG que suporta .lastrowid como SQLite"""
    def __init__(self, cur):
        self._cur = cur
        self.lastrowid = None
        # Tenta obter lastval após INSERT
        try:
            self._cur.execute("SELECT lastval()")
            row = self._cur.fetchone()
            if row:
                self.lastrowid = row.get("lastval") or row.get(0)
        except:
            pass
    def fetchone(self): return self._cur.fetchone()
    def fetchall(self): return self._cur.fetchall()
    def __iter__(self): return iter(self._cur.fetchall())

class _PGConnWrapper:
    """Wrapper de conexão PG que imita interface SQLite"""
    def __init__(self, conn):
        self._conn = conn
    def execute(self, sql, params=None):
        # Converte ? para %s (PostgreSQL não aceita ?)
        pg_sql = sql.replace("?", "%s")
        cur = self._conn.cursor()
        if params:
            cur.execute(pg_sql, params)
        else:
            cur.execute(pg_sql)
        # Se foi INSERT, tenta pegar lastrowid via lastval()
        wrapper = type('_R', (), {
            'lastrowid': None,
            'fetchone': cur.fetchone,
            'fetchall': cur.fetchall
        })()
        if pg_sql.strip().upper().startswith('INSERT'):
            try:
                cur2 = self._conn.cursor()
                cur2.execute("SELECT lastval()")
                row = cur2.fetchone()
                if row:
                    wrapper.lastrowid = row.get("lastval") or list(row.values())[0]
            except:
                pass
        return wrapper
    def commit(self): self._conn.commit()
    def rollback(self): self._conn.rollback()
    def close(self): self._conn.close()
    def cursor(self): return self._conn.cursor()
    def __enter__(self): return self
    def __exit__(self, *a):
        if a[0]: self._conn.rollback()
        else: self._conn.commit()
        return False

def get_db():
    if USE_PG:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
        conn.autocommit = False
        return _PGConnWrapper(conn)
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn

def _exec_pg(conn, sql):
    """Executa bloco SQL no PostgreSQL, ignorando erros de 'já existe'"""
    import psycopg2 as _pg
    cur = conn.cursor()
    for stmt in sql.split(";"):
        stmt = stmt.strip()
        if not stmt or stmt.startswith("--"):
            continue
        try:
            cur.execute(stmt)
            conn.commit()
        except _pg.errors.DuplicateTable:
            conn.rollback()
        except _pg.errors.DuplicateObject:
            conn.rollback()
        except _pg.errors.UniqueViolation:
            conn.rollback()
        except Exception as e:
            conn.rollback()
            logger.warning(f"init_db PG ignorou: {e}")

def _pg_schema():
    """Schema PostgreSQL — sem AUTOINCREMENT, sem PRAGMA, sem SQLite-specific"""
    return """
CREATE TABLE IF NOT EXISTS usuarios (
    id         SERIAL PRIMARY KEY,
    nome       TEXT NOT NULL,
    email      TEXT NOT NULL UNIQUE,
    senha_hash TEXT NOT NULL,
    cargo      TEXT DEFAULT 'voluntario',
    ativo      INTEGER DEFAULT 1,
    criado_em  TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS'),
    ultimo_acesso TEXT DEFAULT NULL
);
CREATE TABLE IF NOT EXISTS cultos (
    id          SERIAL PRIMARY KEY,
    data        TEXT NOT NULL,
    hora        TEXT NOT NULL,
    dia_semana  TEXT NOT NULL,
    periodo     TEXT NOT NULL,
    tipo_culto  TEXT DEFAULT 'Culto Regular',
    responsavel TEXT NOT NULL,
    presentes   INTEGER DEFAULT 0,
    visitantes  INTEGER DEFAULT 0,
    criancas    INTEGER DEFAULT 0,
    observacoes TEXT DEFAULT '',
    usuario_id  INTEGER REFERENCES usuarios(id),
    editado_em  TEXT DEFAULT NULL,
    editado_por TEXT DEFAULT NULL,
    criado_em   TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
CREATE TABLE IF NOT EXISTS cultos_historico (
    id          SERIAL PRIMARY KEY,
    culto_id    INTEGER REFERENCES cultos(id) ON DELETE CASCADE,
    campo       TEXT NOT NULL,
    valor_antes TEXT DEFAULT '',
    valor_depois TEXT DEFAULT '',
    alterado_por TEXT DEFAULT '',
    alterado_em TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
CREATE TABLE IF NOT EXISTS visitantes (
    id          SERIAL PRIMARY KEY,
    culto_id    INTEGER REFERENCES cultos(id) ON DELETE SET NULL,
    nome        TEXT NOT NULL,
    idade       TEXT DEFAULT '',
    telefone    TEXT NOT NULL,
    endereco    TEXT DEFAULT '',
    endereco_padronizado TEXT DEFAULT '',
    cidade      TEXT DEFAULT '',
    bairro      TEXT DEFAULT '',
    cep         TEXT DEFAULT '',
    lat         REAL DEFAULT NULL,
    lng         REAL DEFAULT NULL,
    como_conheceu TEXT DEFAULT '',
    pedido_oracao TEXT DEFAULT '',
    quer_visita INTEGER DEFAULT 0,
    data_visita TEXT DEFAULT '',
    hora_visita TEXT DEFAULT '',
    observacao  TEXT DEFAULT '',
    origem      TEXT DEFAULT 'manual',
    criado_em   TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
CREATE TABLE IF NOT EXISTS checklists (
    id             SERIAL PRIMARY KEY,
    culto_id       INTEGER REFERENCES cultos(id) ON DELETE CASCADE,
    categoria      TEXT NOT NULL,
    item_key       TEXT NOT NULL,
    item_descricao TEXT NOT NULL,
    concluido      INTEGER DEFAULT 0,
    responsavel    TEXT DEFAULT '',
    criado_em      TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
CREATE TABLE IF NOT EXISTS itens_checklist_padrao (
    id        SERIAL PRIMARY KEY,
    categoria TEXT NOT NULL,
    ordem     INTEGER DEFAULT 0,
    descricao TEXT NOT NULL,
    item_key  TEXT NOT NULL UNIQUE
);
CREATE TABLE IF NOT EXISTS estoque (
    id                SERIAL PRIMARY KEY,
    nome              TEXT NOT NULL UNIQUE,
    categoria         TEXT DEFAULT 'Geral',
    quantidade        INTEGER DEFAULT 0,
    quantidade_minima INTEGER DEFAULT 0,
    unidade           TEXT DEFAULT 'unidade',
    descricao         TEXT DEFAULT '',
    fixo              INTEGER DEFAULT 0,
    criado_em         TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS'),
    atualizado_em     TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
CREATE TABLE IF NOT EXISTS grupos_crescimento (
    id        SERIAL PRIMARY KEY,
    nome      TEXT NOT NULL UNIQUE,
    lider     TEXT DEFAULT '',
    telefone_lider TEXT DEFAULT '',
    endereco  TEXT NOT NULL,
    bairro    TEXT DEFAULT '',
    cidade    TEXT DEFAULT 'Alvorada',
    setor     TEXT DEFAULT 'Verde',
    cor_hex   TEXT DEFAULT '#22C55E',
    lat       REAL DEFAULT NULL,
    lng       REAL DEFAULT NULL,
    ativo     INTEGER DEFAULT 1,
    criado_em TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
CREATE TABLE IF NOT EXISTS gc_direcionamentos (
    id             SERIAL PRIMARY KEY,
    visitante_id   INTEGER REFERENCES visitantes(id) ON DELETE SET NULL,
    gc_id          INTEGER REFERENCES grupos_crescimento(id) ON DELETE SET NULL,
    visitante_nome TEXT DEFAULT '',
    gc_nome        TEXT DEFAULT '',
    distancia_km   REAL DEFAULT NULL,
    criado_em      TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
CREATE TABLE IF NOT EXISTS cameras (
    id        SERIAL PRIMARY KEY,
    nome      TEXT NOT NULL,
    url       TEXT NOT NULL,
    local     TEXT DEFAULT '',
    ativa     INTEGER DEFAULT 1,
    criado_em TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
CREATE TABLE IF NOT EXISTS contagem_sessoes (
    id              SERIAL PRIMARY KEY,
    culto_id        INTEGER REFERENCES cultos(id) ON DELETE SET NULL,
    camera_id       INTEGER REFERENCES cameras(id) ON DELETE SET NULL,
    camera_nome     TEXT DEFAULT '',
    iniciado_em     TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS'),
    encerrado_em    TEXT DEFAULT NULL,
    total_entradas  INTEGER DEFAULT 0,
    total_saidas    INTEGER DEFAULT 0,
    pico_simultaneo INTEGER DEFAULT 0,
    status          TEXT DEFAULT 'ativa'
);
CREATE TABLE IF NOT EXISTS contagem_registros (
    id            SERIAL PRIMARY KEY,
    sessao_id     INTEGER REFERENCES contagem_sessoes(id) ON DELETE CASCADE,
    track_id      INTEGER NOT NULL,
    direcao       TEXT NOT NULL,
    confianca     REAL DEFAULT 1.0,
    registrado_em TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
);
"""

def _pg_inserts():
    """Dados iniciais para PostgreSQL"""
    return [
        ("INSERT INTO usuarios (nome,email,senha_hash,cargo) VALUES (%s,%s,%s,%s) ON CONFLICT (email) DO NOTHING",
         ('Administrador','admin@igrejaaba.com','e86f78a8a3caf0b60d8e74e5942aa6d86dc150cd3c03338aef25b7d2d7e3acc7','admin')),
    ]

_CHECKLIST_ITEMS = [
    ('antes',1,'Verificar se tem copos no bebedouro','ant_copos'),
    ('antes',2,'Equipe do estacionamento: usar coletes, distribuir cones','ant_estac'),
    ('antes',3,'Ligar os ar-condicionados em dias de calor','ant_ar'),
    ('antes',4,'Estacionamento organizado','ant_estac2'),
    ('antes',5,'Usar crachá','ant_cracha'),
    ('antes',6,'Varrer a frente da igreja e a área da cantina','ant_varrer'),
    ('mesa_entrada',1,'Envelopes de dízimos e oferta','mesa_envelopes'),
    ('mesa_entrada',2,'Fichas "Quero ser membro de um GC"','mesa_fichas_gc'),
    ('mesa_entrada',3,'Fichas "Preciso de oração"','mesa_fichas_oracao'),
    ('mesa_entrada',4,'Organizar os vales presentes dos visitantes','mesa_vales'),
    ('banheiro',1,'Verificar papel higiênico','ban_papel_hig'),
    ('banheiro',2,'Verificar papel toalha','ban_papel_toalha'),
    ('banheiro',3,'Verificar sabonete líquido','ban_sabonete'),
    ('banheiro',4,'Verificar lixeiras','ban_lixeiras'),
    ('durante',1,'Distribuir envelopes na oferta','dur_envelopes'),
    ('durante',2,'Levar água ao ministrador','dur_agua'),
    ('durante',3,'Atenção nas situações diversas','dur_atencao'),
    ('durante',4,'Contagem de presentes, visitantes e crianças','dur_contagem'),
    ('durante',5,'Entregar vale presente','dur_vale'),
    ('final',1,'Retirar lixo','fin_lixo'),
    ('final',2,'Organizar cadeiras','fin_cadeiras'),
    ('final',3,'Desligar ar-condicionado','fin_ar'),
    ('final',4,'Verificar se todas as luzes estão apagadas','fin_luzes'),
    ('final',5,'Verificar as torneiras dos banheiros','fin_torneiras'),
    ('final',6,'Fechar portas','fin_portas'),
    ('final',7,'Acionar alarme','fin_alarme'),
    ('final',8,'Recolher cones e placas','fin_cones'),
]

_GCS = [
    # (nome, lider, endereco, bairro, cidade, setor, cor_hex, lat, lng)
    ('GC Conectados - Intersul','','Av. Borges de Medeiros, 196','Intersul','Alvorada','Amarelo','#EAB308',-30.0195,-51.072),
    ('GC Conectados - Jardim Algarve','Gomes e Marilu','Rua Hermínio Machado, 475','Jardim Algarve','Alvorada','Amarelo','#EAB308',-30.0301,-51.0826),
    ('GC Conectados - Porto Verde','','Rua Beija-flores, 371','Porto Verde','Alvorada','Amarelo','#EAB308',-29.9745,-51.0823),
    ('GC Corujas','Dinho e Andressa','Rua Corujas, 552','Porto Verde','Alvorada','Azul','#3B82F6',-30.0404,-51.0751),
    ('GC Master Fé','Eduardo e Vanessa','Rua Gonçalves de Magalhães, 806','Jardim Porto Alegre','Alvorada','Azul','#3B82F6',-30.0243,-51.0766),
    ('GC Caraá','Nubia e Matheus','Rua Hermínio Machado, 574','Rio dos Sinos','Caraá','Laranja','#F97316',-30.031,-51.0827),
    ('GC Luz do Mundo','Adriel e Paola','Rua Alameda, 97','Jardim Algarve','Alvorada','Laranja','#F97316',-30.0287,-51.0853),
    ('GC Maranata','Oriton e Eliane','Rua Pedro Claudio Monassa, 380','Jardim Algarve','Alvorada','Roxo','#A855F7',-30.0292,-51.0813),
    ('GC Resgate da Cruz','Regis e Gilda','Av. Elmira Pereira Silveira, 327','Jardim Algarve','Alvorada','Roxo','#A855F7',-30.0309,-51.0838),
    ('GC Infinito e Amém','Gabriel e Bruna','Rua Cento e Trinta e Nove, 84','Jardim Algarve','Alvorada','Verde','#22C55E',-30.0344,-51.0859),
    ('GC Farol da Lagoa','Vanessa e Lucas','Av. Borges de Medeiros, 196','Intersul','Alvorada','Vermelho','#EF4444',-30.0199,-51.0719),
    ('GC Manálovers','Juliana e Luiz','Rua Flaviano Morais Monroe, 556','Jardim Algarve','Alvorada','Vermelho','#EF4444',-30.0324,-51.0872),
    ('GC Palavra Viva','','Rua Trinta e Quatro, 318','Jardim Algarve','Alvorada','Vermelho','#EF4444',-30.0248,-51.0811)
]

_ESTOQUE = [
    ('Cálices de Santa Ceia — Individuais','Santa Ceia',0,50,'unidade','Cálices descartáveis individuais'),
]

def init_db():
    if USE_PG:
        _init_pg()
    else:
        _init_sqlite()


def _pg_migrations(conn):
    """Adiciona colunas/tabelas que podem faltar em bancos antigos (safe ALTER TABLE)"""
    migrations = [
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS ultimo_acesso TEXT DEFAULT NULL",
        """CREATE TABLE IF NOT EXISTS relatorios_gc (
            id                SERIAL PRIMARY KEY,
            gc_id             INTEGER REFERENCES grupos_crescimento(id) ON DELETE SET NULL,
            gc_nome           TEXT NOT NULL,
            lider_nome        TEXT NOT NULL,
            anfitriao         TEXT DEFAULT '',
            dia               TEXT NOT NULL,
            membros_presentes INTEGER DEFAULT 0,
            visitantes        INTEGER DEFAULT 0,
            lider_treinamento INTEGER DEFAULT 0,
            nome_lider_trein  TEXT DEFAULT '',
            observacoes       TEXT DEFAULT '',
            criado_em         TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')
        )""",
        "ALTER TABLE grupos_crescimento ADD COLUMN IF NOT EXISTS lider TEXT DEFAULT ''",
        """CREATE TABLE IF NOT EXISTS voluntarios (
            id SERIAL PRIMARY KEY, nome TEXT NOT NULL, telefone TEXT NOT NULL,
            departamentos TEXT DEFAULT '',
            departamento_id INTEGER REFERENCES departamentos(id) ON DELETE SET NULL,
            usuario_id INTEGER REFERENCES usuarios(id) ON DELETE SET NULL,
            ativo INTEGER DEFAULT 1,
            criado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'))""",
        """CREATE TABLE IF NOT EXISTS escala_publicacoes (
            id SERIAL PRIMARY KEY, mes TEXT NOT NULL UNIQUE,
            status TEXT DEFAULT 'rascunho', publicado_em TEXT DEFAULT NULL,
            publicado_por TEXT DEFAULT '')""",
        """CREATE TABLE IF NOT EXISTS escala_confirmacoes (
            id SERIAL PRIMARY KEY,
            voluntario_id INTEGER REFERENCES voluntarios(id) ON DELETE CASCADE,
            voluntario_nome TEXT DEFAULT '', culto_data TEXT NOT NULL,
            culto_periodo TEXT NOT NULL, departamento TEXT NOT NULL,
            status TEXT DEFAULT 'pendente', sugestao_troca TEXT DEFAULT '',
            token TEXT NOT NULL UNIQUE, notificado_em TEXT DEFAULT NULL,
            respondido_em TEXT DEFAULT NULL,
            criado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'))""",
        """CREATE TABLE IF NOT EXISTS departamentos (
            id SERIAL PRIMARY KEY, nome TEXT NOT NULL UNIQUE,
            icone TEXT DEFAULT '👥', ordem INTEGER DEFAULT 0, ativo INTEGER DEFAULT 1)""",
        """CREATE TABLE IF NOT EXISTS escala_itens (
            id SERIAL PRIMARY KEY,
            departamento_id INTEGER REFERENCES departamentos(id) ON DELETE CASCADE,
            culto_data TEXT NOT NULL, culto_periodo TEXT NOT NULL,
            responsavel TEXT DEFAULT '', observacao TEXT DEFAULT '',
            criado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'),
            atualizado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'))""",
        "ALTER TABLE grupos_crescimento ADD COLUMN IF NOT EXISTS telefone_lider TEXT DEFAULT ''",
        "ALTER TABLE cultos ADD COLUMN IF NOT EXISTS tipo_culto TEXT DEFAULT 'Culto Regular'",
        "ALTER TABLE cultos ADD COLUMN IF NOT EXISTS editado_em TEXT DEFAULT NULL",
        "ALTER TABLE cultos ADD COLUMN IF NOT EXISTS editado_por TEXT DEFAULT NULL",
        "ALTER TABLE visitantes ADD COLUMN IF NOT EXISTS endereco_padronizado TEXT DEFAULT ''",
        "ALTER TABLE visitantes ADD COLUMN IF NOT EXISTS lat REAL DEFAULT NULL",
        "ALTER TABLE visitantes ADD COLUMN IF NOT EXISTS lng REAL DEFAULT NULL",
        "ALTER TABLE estoque ADD COLUMN IF NOT EXISTS atualizado_em TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS criado_em TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')",
        "ALTER TABLE cultos ADD COLUMN IF NOT EXISTS tipo_culto TEXT DEFAULT 'Culto Regular'",
        "ALTER TABLE cultos ADD COLUMN IF NOT EXISTS editado_em TEXT DEFAULT NULL",
        "ALTER TABLE cultos ADD COLUMN IF NOT EXISTS editado_por TEXT DEFAULT NULL",
        "ALTER TABLE visitantes ADD COLUMN IF NOT EXISTS endereco_padronizado TEXT DEFAULT ''",
        "ALTER TABLE visitantes ADD COLUMN IF NOT EXISTS lat REAL DEFAULT NULL",
        "ALTER TABLE visitantes ADD COLUMN IF NOT EXISTS lng REAL DEFAULT NULL",
        "ALTER TABLE grupos_crescimento ADD COLUMN IF NOT EXISTS lider TEXT DEFAULT ''",
        """CREATE TABLE IF NOT EXISTS voluntarios (
            id SERIAL PRIMARY KEY, nome TEXT NOT NULL, telefone TEXT NOT NULL,
            departamentos TEXT DEFAULT '', ativo INTEGER DEFAULT 1,
            criado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'))""",
        """CREATE TABLE IF NOT EXISTS escala_publicacoes (
            id SERIAL PRIMARY KEY, mes TEXT NOT NULL UNIQUE,
            status TEXT DEFAULT 'rascunho', publicado_em TEXT DEFAULT NULL,
            publicado_por TEXT DEFAULT '')""",
        """CREATE TABLE IF NOT EXISTS escala_confirmacoes (
            id SERIAL PRIMARY KEY,
            voluntario_id INTEGER REFERENCES voluntarios(id) ON DELETE CASCADE,
            voluntario_nome TEXT DEFAULT '', culto_data TEXT NOT NULL,
            culto_periodo TEXT NOT NULL, departamento TEXT NOT NULL,
            status TEXT DEFAULT 'pendente', sugestao_troca TEXT DEFAULT '',
            token TEXT NOT NULL UNIQUE, notificado_em TEXT DEFAULT NULL,
            respondido_em TEXT DEFAULT NULL,
            criado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'))""",
        """CREATE TABLE IF NOT EXISTS departamentos (
            id SERIAL PRIMARY KEY, nome TEXT NOT NULL UNIQUE,
            icone TEXT DEFAULT '👥', ordem INTEGER DEFAULT 0, ativo INTEGER DEFAULT 1)""",
        """CREATE TABLE IF NOT EXISTS escala_itens (
            id SERIAL PRIMARY KEY,
            departamento_id INTEGER REFERENCES departamentos(id) ON DELETE CASCADE,
            culto_data TEXT NOT NULL, culto_periodo TEXT NOT NULL,
            responsavel TEXT DEFAULT '', observacao TEXT DEFAULT '',
            criado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'),
            atualizado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'))""",
        "ALTER TABLE grupos_crescimento ADD COLUMN IF NOT EXISTS telefone_lider TEXT DEFAULT ''",
        "ALTER TABLE estoque ADD COLUMN IF NOT EXISTS atualizado_em TEXT DEFAULT to_char(now(), 'YYYY-MM-DD HH24:MI:SS')",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS departamento_id INTEGER REFERENCES departamentos(id) ON DELETE SET NULL",
        "ALTER TABLE voluntarios ADD COLUMN IF NOT EXISTS departamento_id INTEGER REFERENCES departamentos(id) ON DELETE SET NULL",
        "ALTER TABLE voluntarios ADD COLUMN IF NOT EXISTS usuario_id INTEGER REFERENCES usuarios(id) ON DELETE SET NULL",
        "ALTER TABLE relatorios_gc ADD COLUMN IF NOT EXISTS campos_livres TEXT DEFAULT ''",
        "ALTER TABLE grupos_crescimento ADD COLUMN IF NOT EXISTS gc_pai_id INTEGER REFERENCES grupos_crescimento(id) ON DELETE SET NULL",
        "ALTER TABLE grupos_crescimento ADD COLUMN IF NOT EXISTS supervisor TEXT DEFAULT ''",
        "ALTER TABLE relatorios_gc ADD COLUMN IF NOT EXISTS valor_oferta REAL DEFAULT 0",
        "ALTER TABLE relatorios_gc ADD COLUMN IF NOT EXISTS quilos_arrecadados REAL DEFAULT 0",
        """CREATE TABLE IF NOT EXISTS gc_integrantes (
            id SERIAL PRIMARY KEY,
            gc_id INTEGER REFERENCES grupos_crescimento(id) ON DELETE CASCADE,
            nome TEXT NOT NULL, telefone TEXT DEFAULT '',
            ativo INTEGER DEFAULT 1,
            criado_em TEXT DEFAULT to_char(now(),'YYYY-MM-DD HH24:MI:SS'))""",
    ]
    cur = conn.cursor()
    for sql in migrations:
        try:
            cur.execute(sql)
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.warning(f"Migration ignorada: {e}")

def _init_pg():
    """Inicializa banco PostgreSQL"""
    try:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
        conn.autocommit = False
        _exec_pg(conn, _pg_schema())
        # Migrations seguras — adiciona colunas que podem não existir em bancos antigos
        _pg_migrations(conn)
        # Limpa duplicatas antes de criar índice único
        cur = conn.cursor()
        try:
            cur.execute("""
                DELETE FROM grupos_crescimento WHERE id NOT IN (
                    SELECT MIN(id) FROM grupos_crescimento GROUP BY nome
                )
            """)
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.warning(f"Limpeza duplicatas GC: {e}")

        # Cria índice único no nome do GC (se não existir)
        try:
            cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_gc_nome ON grupos_crescimento(nome)")
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.warning(f"Índice GC: {e}")

        # Admin padrão
        cur.execute(
            "INSERT INTO usuarios (nome,email,senha_hash,cargo) VALUES (%s,%s,%s,%s) ON CONFLICT (email) DO NOTHING",
            ('Administrador','admin@igrejaaba.com','e86f78a8a3caf0b60d8e74e5942aa6d86dc150cd3c03338aef25b7d2d7e3acc7','admin')
        )
        # Checklist padrão
        for item in _CHECKLIST_ITEMS:
            cur.execute(
                "INSERT INTO itens_checklist_padrao (categoria,ordem,descricao,item_key) VALUES (%s,%s,%s,%s) ON CONFLICT (item_key) DO NOTHING",
                item
            )
        # GCs
        for gc in _GCS:
            try:
                cur.execute(
                    """INSERT INTO grupos_crescimento (nome,lider,endereco,bairro,cidade,setor,cor_hex,lat,lng)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT (nome) DO UPDATE SET
                         lider=EXCLUDED.lider, endereco=EXCLUDED.endereco,
                         bairro=EXCLUDED.bairro, cidade=EXCLUDED.cidade,
                         setor=EXCLUDED.setor, cor_hex=EXCLUDED.cor_hex,
                         lat=EXCLUDED.lat, lng=EXCLUDED.lng""",
                    gc
                )
            except Exception as e:
                conn.rollback()
                logger.warning(f"Insert GC {gc[0]}: {e}")
        # Estoque
        for est in _ESTOQUE:
            cur.execute(
                "INSERT INTO estoque (nome,categoria,quantidade,quantidade_minima,unidade,descricao,fixo) VALUES (%s,%s,%s,%s,%s,%s,0) ON CONFLICT (nome) DO NOTHING",
                est
            )
        # Câmera padrão

        # Atualiza coordenadas fixas dos GCs conhecidos
        _gc_coords = {
            'GC Conectados - Intersul': (-29.9758, -51.0812),
            'GC Conectados - Jardim Algarve': (-30.0301, -51.0826),
            'GC Conectados - Porto Verde': (-29.9745, -51.0823),
            'GC Corujas': (-30.0404, -51.0751),
            'GC Master Fe': (-30.0243, -51.0766),
            'GC Caraa': (-30.031, -51.0827),
            'GC Luz do Mundo': (-30.0287, -51.0853),
            'GC Maranata': (-30.0292, -51.0813),
            'GC Resgate da Cruz': (-30.0309, -51.0838),
            'GC Infinito e Amem': (-30.0344, -51.0859),
            'GC Farol da Lagoa': (-30.0199, -51.0719),
            'GC Manalovrs': (-30.0324, -51.0872),
            'GC Palavra Viva': (-30.0248, -51.0811)
        }
        for nome, (lat, lng) in _gc_coords.items():
            try:
                cur.execute(
                    "UPDATE grupos_crescimento SET lat=%s, lng=%s WHERE nome=%s",
                    (lat, lng, nome)
                )
                conn.commit()
            except Exception as e:
                conn.rollback()

        # Seed departamentos
        _departs = [('Recepção', '🤝', 1), ('Estacionamento', '🚗', 2), ('Mídia / Stores', '📱', 3), ('Som e Iluminação', '🎛️', 4), ('Louvor', '🎵', 5), ('Abertura', '🎤', 6), ('Oração', '🙏', 7), ('Ministração', '✝️', 8), ('Oferta', '💛', 9), ('Encerramento', '🏁', 10), ('Kids', '👶', 11), ('Pré-Teens', '🧒', 12), ('Cantina', '☕', 13)]
        for dep in _departs:
            try:
                cur.execute("INSERT INTO departamentos(nome,icone,ordem) VALUES(%s,%s,%s) ON CONFLICT(nome) DO NOTHING", dep)
                conn.commit()
            except Exception as e:
                conn.rollback()
        # Câmera: só insere se não existir nenhuma
        cur.execute("SELECT COUNT(*) as n FROM cameras")
        row = cur.fetchone()
        n_cam = row["n"] if row and isinstance(row, dict) else (row[0] if row else 0)
        if n_cam == 0:
            cur.execute(
                "INSERT INTO cameras (nome,url,local) VALUES (%s,%s,%s)",
                ('Câmera Principal','0','Entrada Principal')
            )
        conn.commit()
        conn.close()
        logger.info("PostgreSQL inicializado com sucesso!")
    except Exception as e:
        logger.error(f"Erro init PostgreSQL: {e}")
        raise

def _init_sqlite():
    """Inicializa banco SQLite"""
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    with open(SQL_PATH,"r",encoding="utf-8") as sf:
        conn.executescript(sf.read())
    # Migrations seguras para bancos existentes
    _migrations_sqlite = [
        "ALTER TABLE usuarios ADD COLUMN departamento_id INTEGER REFERENCES departamentos(id) ON DELETE SET NULL",
        "ALTER TABLE usuarios ADD COLUMN ultimo_acesso TEXT DEFAULT NULL",
        "ALTER TABLE voluntarios ADD COLUMN departamento_id INTEGER REFERENCES departamentos(id) ON DELETE SET NULL",
        "ALTER TABLE voluntarios ADD COLUMN usuario_id INTEGER REFERENCES usuarios(id) ON DELETE SET NULL",
        "ALTER TABLE relatorios_gc ADD COLUMN campos_livres TEXT DEFAULT ''",
        "ALTER TABLE grupos_crescimento ADD COLUMN gc_pai_id INTEGER REFERENCES grupos_crescimento(id) ON DELETE SET NULL",
        "ALTER TABLE grupos_crescimento ADD COLUMN supervisor TEXT DEFAULT ''",
        "ALTER TABLE relatorios_gc ADD COLUMN valor_oferta REAL DEFAULT 0",
        "ALTER TABLE relatorios_gc ADD COLUMN quilos_arrecadados REAL DEFAULT 0",
        """CREATE TABLE IF NOT EXISTS gc_integrantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gc_id INTEGER REFERENCES grupos_crescimento(id) ON DELETE CASCADE,
            nome TEXT NOT NULL, telefone TEXT DEFAULT '',
            ativo INTEGER DEFAULT 1,
            criado_em TEXT DEFAULT (datetime('now','localtime')))""",
    ]
    for sql in _migrations_sqlite:
        try:
            conn.execute(sql)
            conn.commit()
        except Exception:
            conn.rollback()
    conn.execute("DELETE FROM estoque WHERE id NOT IN (SELECT MIN(id) FROM estoque GROUP BY nome)")
    try:
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_estoque_nome ON estoque(nome)")
    except: pass
    conn.commit()
    conn.close()
    logger.info(f"SQLite inicializado: {DB_PATH}")




def get_lastid(conn, table=None):
    """Obtém o último ID inserido (PG usa lastval, SQLite usa lastrowid de outra forma)"""
    if USE_PG:
        try:
            cur2 = conn.cursor()
            cur2.execute("SELECT lastval()")
            row = cur2.fetchone()
            return row["lastval"] if row else None
        except:
            return None
    return None  # SQLite usa cur.lastrowid diretamente

def executar_insert(conn, sql, params):
    """Executa INSERT e retorna o ID inserido (funciona em PG e SQLite)"""
    if USE_PG:
        # Adiciona RETURNING id se não tiver
        sql_ret = sql.rstrip().rstrip(")").rstrip()
        if "RETURNING" not in sql.upper():
            # Precisamos usar cursor diretamente no PG
            cur = conn.cursor()
            cur.execute(qmark(sql), params)
            # Tenta obter id via lastval
            cur.execute("SELECT lastval()")
            row = cur.fetchone()
            return row["lastval"] if row else None
        else:
            cur = conn.cursor()
            cur.execute(qmark(sql), params)
            row = cur.fetchone()
            return row["id"] if row else None
    else:
        cur = conn.execute(sql, params)
        return cur.lastrowid

def ph(n=1):
    """Retorna placeholders corretos: %s para PG, ? para SQLite"""
    if USE_PG:
        return ",".join(["%s"]*n) if n>1 else "%s"
    return ",".join(["?"]*n) if n>1 else "?"

def qmark(sql):
    """Converte ? para %s quando usando PostgreSQL"""
    if USE_PG:
        return sql.replace("?", "%s")
    return sql

def lastid(conn, cur_or_result, table=None):
    """Retorna último ID inserido"""
    if USE_PG:
        r = cur_or_result.fetchone()
        return r["id"] if r else None
    return cur_or_result.lastrowid

def hs(s): return hashlib.sha256(s.encode()).hexdigest()

# ── Helpers ───────────────────────────────────────────────────
DIAS = {0:"Segunda-feira",1:"Terça-feira",2:"Quarta-feira",
        3:"Quinta-feira",4:"Sexta-feira",5:"Sábado",6:"Domingo"}

def dia_pt(s):
    try: return DIAS[datetime.strptime(s,"%Y-%m-%d").weekday()]
    except: return ""

def br(s):
    try: return datetime.strptime(s,"%Y-%m-%d").strftime("%d/%m/%Y")
    except: return s or ""

def _eh_domingo(data_str):
    """Retorna True se a data cai em dia de culto: sexta(4), sábado(5) ou domingo(6)"""
    try:
        return datetime.strptime(str(data_str)[:10], "%Y-%m-%d").weekday() in (4, 5, 6)
    except Exception:
        return False

def _periodos_do_dia(data_str):
    """Retorna os períodos válidos para o dia: Sex/Sáb=Noite, Dom=Manhã+Noite"""
    try:
        wd = datetime.strptime(str(data_str)[:10], "%Y-%m-%d").weekday()
    except Exception:
        return []
    if wd == 6:        # domingo
        return ["Manhã", "Noite"]
    if wd in (4, 5):   # sexta, sábado
        return ["Noite"]
    return []

def get_base():
    b = os.environ.get("BASE_URL","").rstrip("/")
    if b: return b
    proto = request.headers.get("X-Forwarded-Proto", request.scheme)
    host  = request.headers.get("X-Forwarded-Host",  request.host)
    return f"{proto}://{host}"

def haversine(la1,lo1,la2,lo2):
    R=6371; r=math.pi/180
    d1=(la2-la1)*r; d2=(lo2-lo1)*r
    a=math.sin(d1/2)**2+math.cos(la1*r)*math.cos(la2*r)*math.sin(d2/2)**2
    return R*2*math.asin(math.sqrt(max(0,a)))


# ── Geocode melhorado com Nominatim ──────────────────────────
GOOGLE_MAPS_KEY = os.environ.get("GOOGLE_MAPS_KEY", "AIzaSyC6MJNveTAoroPrfDbBMqFl3jp-fEBfBwI")

def geocode_smart(query, cidade_fallback="Alvorada"):
    """Geocode via Google Maps API — preciso e confiável."""
    import urllib.request as _ur, json as _js
    q = (query or "").strip()
    if not q: return None, None, ""
    cidade = cidade_fallback or "Alvorada"
    # Monta query completa para o Google
    full_q = f"{q}, {cidade}, RS, Brasil"
    url = (f"https://maps.googleapis.com/maps/api/geocode/json"
           f"?address={urllib.parse.quote(full_q)}"
           f"&key={GOOGLE_MAPS_KEY}"
           f"&language=pt-BR"
           f"&region=br")
    try:
        req = _ur.Request(url)
        with _ur.urlopen(req, timeout=10) as resp:
            data = _js.loads(resp.read())
        if data.get("status") == "OK" and data.get("results"):
            r = data["results"][0]
            loc = r["geometry"]["location"]
            return float(loc["lat"]), float(loc["lng"]), r.get("formatted_address","")
        else:
            logger.warning(f"Google Geocode status: {data.get('status')} para: {full_q}")
    except Exception as e:
        logger.warning(f"Google Geocode erro: {e}")
    return None, None, ""

# ── Auth decorators ───────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def dec(*a,**k):
        if "usuario_id" not in session:
            return jsonify({"erro":"Não autenticado"}),401
        return f(*a,**k)
    return dec

def role_required(*roles):
    """Exige que o usuário tenha um dos cargos especificados"""
    def decorator(f):
        @wraps(f)
        def dec(*a,**k):
            if "usuario_id" not in session:
                return jsonify({"erro":"Não autenticado"}),401
            cargo = session.get("usuario_cargo","voluntario")
            if cargo not in roles:
                return jsonify({"erro":"Sem permissão para esta ação"}),403
            return f(*a,**k)
        return dec
    return decorator

def can_edit():
    """Líderes e admins podem editar"""
    return session.get("usuario_cargo","voluntario") in ("lider","admin")

def is_admin():
    return session.get("usuario_cargo") == "admin"

def get_depto_usuario():
    """Retorna departamento_id do usuário logado. None = acesso total (admin/lider)"""
    cargo = session.get("usuario_cargo","")
    if cargo == "lider_depto":
        return session.get("usuario_departamento_id")
    return None  # admin e lider veem tudo



# ── Erros ─────────────────────────────────────────────────────
@app.errorhandler(500)
def e500(e): return jsonify({"erro":"Erro interno","d":str(e)}),500
@app.errorhandler(404)
def e404(e): return jsonify({"erro":"Não encontrado"}),404

# ═══════════════════════════════════════════════════════════════
# PÁGINAS
# ═══════════════════════════════════════════════════════════════
@app.route("/")
def index(): return render_template("login.html")

@app.route("/app")
def app_main():
    if "usuario_id" not in session: return redirect("/")
    return render_template("dashboard.html")

@app.route("/formulario")
def formulario():
    return render_template("visitante_form.html", culto_id=request.args.get("culto_id",""))

@app.route("/static/sw.js")
def service_worker():
    from flask import send_from_directory, Response
    sw_path = os.path.join(BASE_DIR, "static", "sw.js")
    with open(sw_path) as f:
        content = f.read()
    return Response(content, mimetype="application/javascript",
                    headers={"Service-Worker-Allowed": "/"})

@app.route("/health")
def health():
    try:
        with get_db() as conn: conn.execute("SELECT 1")
        return jsonify({"status":"ok","time":datetime.now().isoformat()})
    except Exception as e:
        return jsonify({"status":"error","e":str(e)}),500

# ═══════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════
@app.route("/api/login", methods=["POST"])
def login():
    try:
        d     = request.get_json(force=True) or {}
        email = d.get("email","").strip().lower()
        senha = d.get("senha","")
        if not email or not senha:
            return jsonify({"erro":"Preencha e-mail e senha"}),400
        with get_db() as conn:
            u = conn.execute(
                qmark("SELECT * FROM usuarios WHERE email=? AND ativo=1"),(email,)
            ).fetchone()
        if not u:
            return jsonify({"erro":"Usuário não encontrado ou inativo"}),401
        if u["senha_hash"] != hs(senha):
            return jsonify({"erro":"Senha incorreta"}),401
        # Acesso seguro a departamento_id (Row não tem .get())
        try:
            dep_id = u["departamento_id"]
        except (KeyError, IndexError):
            dep_id = None
        session.permanent = True
        session["usuario_id"]    = u["id"]
        session["usuario_nome"]  = u["nome"]
        session["usuario_cargo"] = u["cargo"]
        session["usuario_departamento_id"] = dep_id
        # Define redirect baseado no cargo
        redirect_map = {
            "voluntario_escala": "/minha-escala",
            "lider_depto":       "/escala",
        }
        redirect_url = redirect_map.get(u["cargo"], "/app")
        # Atualiza último acesso
        with get_db() as conn:
            try:
                conn.execute(qmark("UPDATE usuarios SET ultimo_acesso=? WHERE id=?"),
                             (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), u["id"]))
                conn.commit()
            except Exception as _e:
                logger.warning(f"Nao salvou ultimo_acesso: {_e}")
                try: conn.rollback()
                except: pass
        return jsonify({"ok":True,"nome":u["nome"],"cargo":u["cargo"],"redirect":redirect_url})
    except Exception as e:
        logger.error(f"Login error: {e}")
        return jsonify({"erro":str(e)}),500

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok":True})

@app.route("/api/me")
def me():
    if "usuario_id" not in session:
        return jsonify({"autenticado":False})
    depto_id = session.get("usuario_departamento_id")
    depto_nome = ""
    if depto_id:
        try:
            with get_db() as conn:
                dep = conn.execute(qmark("SELECT nome,icone FROM departamentos WHERE id=?"), (depto_id,)).fetchone()
                if dep: depto_nome = dep["icone"] + " " + dep["nome"]
        except: pass
    # Busca voluntario_id vinculado ao usuário (se for voluntario_escala)
    vol_id = None
    if session.get("usuario_cargo") == "voluntario_escala":
        try:
            with get_db() as conn:
                vol = conn.execute(
                    qmark("SELECT id FROM voluntarios WHERE usuario_id=? AND ativo=1"),
                    (session["usuario_id"],)
                ).fetchone()
                if vol: vol_id = vol["id"]
        except: pass

    return jsonify({
        "autenticado":      True,
        "id":               session["usuario_id"],
        "nome":             session["usuario_nome"],
        "cargo":            session["usuario_cargo"],
        "departamento_id":  depto_id,
        "departamento_nome":depto_nome,
        "voluntario_id":    vol_id
    })

# ═══════════════════════════════════════════════════════════════
# USUÁRIOS — apenas admin
# ═══════════════════════════════════════════════════════════════
@app.route("/api/usuarios", methods=["GET"])
@login_required
def listar_usuarios():
    with get_db() as conn:
        rows = conn.execute(
            """SELECT u.id,u.nome,u.email,u.cargo,u.ativo,u.criado_em,u.departamento_id,
               COALESCE(d.icone||' '||d.nome,'') as departamento_nome
               FROM usuarios u
               LEFT JOIN departamentos d ON d.id=u.departamento_id
               ORDER BY u.nome"""
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/usuarios", methods=["POST"])
@role_required("admin")
def criar_usuario():
    d     = request.get_json(force=True) or {}
    nome  = d.get("nome","").strip()
    email = d.get("email","").strip().lower()
    senha = d.get("senha","")
    cargo = d.get("cargo","voluntario")
    conf  = d.get("confirmar_senha","")
    if not nome or not email or not senha:
        return jsonify({"erro":"Nome, e-mail e senha são obrigatórios"}),400
    if len(senha) < 8:
        return jsonify({"erro":"Senha deve ter no mínimo 8 caracteres"}),400
    if senha != conf:
        return jsonify({"erro":"As senhas não coincidem"}),400
    if not re.search(r'[A-Z]', senha):
        return jsonify({"erro":"Senha deve ter ao menos uma letra maiúscula"}),400
    if not re.search(r'[0-9]', senha):
        return jsonify({"erro":"Senha deve ter ao menos um número"}),400
    if cargo not in ("voluntario","lider","admin","lider_depto","voluntario_escala"):
        return jsonify({"erro":"Cargo inválido"}),400
    # Departamento obrigatório para líder de departamento
    depto_id = d.get("departamento_id")
    if cargo == "lider_depto" and not depto_id:
        return jsonify({"erro":"Selecione o departamento do líder."}),400
    try:
        with get_db() as conn:
            conn.execute(
                qmark("INSERT INTO usuarios(nome,email,senha_hash,cargo,departamento_id) VALUES(?,?,?,?,?)"),
                (nome, email, hs(senha), cargo, depto_id if cargo=="lider_depto" else None)
            )
            conn.commit()
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"erro":"E-mail já cadastrado" if "UNIQUE" in str(e) else str(e)}),400

@app.route("/api/usuarios/<int:uid>", methods=["PUT"])
@login_required
def editar_usuario(uid):
    # Voluntário só pode mudar a própria senha
    cargo_atual = session.get("usuario_cargo","voluntario")
    meu_id = session["usuario_id"]
    d = request.get_json(force=True) or {}

    if cargo_atual == "voluntario" and meu_id != uid:
        return jsonify({"erro":"Sem permissão"}),403

    with get_db() as conn:
        alvo = conn.execute(qmark("SELECT * FROM usuarios WHERE id=?"), (uid,)).fetchone()
        if not alvo: return jsonify({"erro":"Usuário não encontrado"}),404

        # Ninguém (nem admin) pode mudar cargo de outro admin sem ser admin
        if "cargo" in d and cargo_atual != "admin":
            return jsonify({"erro":"Apenas admins alteram cargos"}),403

        # Não pode excluir/desativar o próprio admin
        if "ativo" in d and uid == meu_id:
            return jsonify({"erro":"Não pode desativar a si mesmo"}),400

        if "nova_senha" in d and d["nova_senha"]:
            nova = d["nova_senha"]
            conf = d.get("confirmar_nova_senha","")
            if len(nova) < 8:
                return jsonify({"erro":"Senha mínima de 8 caracteres"}),400
            if nova != conf:
                return jsonify({"erro":"As senhas não coincidem"}),400
            if not re.search(r'[A-Z]', nova):
                return jsonify({"erro":"Senha deve ter ao menos uma letra maiúscula"}),400
            if not re.search(r'[0-9]', nova):
                return jsonify({"erro":"Senha deve ter ao menos um número"}),400
            conn.execute(qmark("UPDATE usuarios SET senha_hash=? WHERE id=?"), (hs(nova),uid))

        if "nome" in d:
            conn.execute(qmark("UPDATE usuarios SET nome=? WHERE id=?"), (d["nome"],uid))
        if "cargo" in d and cargo_atual == "admin":
            conn.execute(qmark("UPDATE usuarios SET cargo=? WHERE id=?"), (d["cargo"],uid))
        if "ativo" in d and cargo_atual == "admin":
            conn.execute(qmark("UPDATE usuarios SET ativo=? WHERE id=?"), (int(d["ativo"]),uid))
        conn.commit()

    return jsonify({"ok":True})

@app.route("/api/usuarios/<int:uid>", methods=["DELETE"])
@role_required("admin")
def deletar_usuario(uid):
    if uid == session["usuario_id"]:
        return jsonify({"erro":"Não pode excluir a si mesmo"}),400
    with get_db() as conn:
        alvo = conn.execute(qmark("SELECT cargo FROM usuarios WHERE id=?"), (uid,)).fetchone()
        if alvo and alvo["cargo"] == "admin":
            row = conn.execute(qmark("SELECT COUNT(*) as total FROM usuarios WHERE cargo=? AND ativo=1"), ("admin",)).fetchone()
            total_admins = row["total"] if row else 0
            if total_admins <= 1:
                return jsonify({"erro":"Não é possível remover o único administrador"}),400
        conn.execute(qmark("DELETE FROM usuarios WHERE id=?"), (uid,)); conn.commit()
    return jsonify({"ok":True})

# ═══════════════════════════════════════════════════════════════
# CULTOS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/cultos", methods=["GET"])
@login_required
def listar_cultos():
    ini = request.args.get("data_ini","")
    fim = request.args.get("data_fim","")
    per = request.args.get("periodo","")
    tpc = request.args.get("tipo_culto","")
    sql = "SELECT id,data,hora,dia_semana,periodo,COALESCE(tipo_culto,'Culto Regular') as tipo_culto,responsavel,presentes,visitantes,criancas,observacoes,criado_em,editado_em,editado_por FROM cultos WHERE 1=1"
    p   = []
    if ini: sql+=" AND data>=?"; p.append(ini)
    if fim: sql+=" AND data<=?"; p.append(fim)
    if per: sql+=" AND periodo=?"; p.append(per)
    if tpc: sql+=" AND tipo_culto=?"; p.append(tpc)
    sql += " ORDER BY data DESC,hora DESC"
    with get_db() as conn:
        rows = conn.execute(sql,p).fetchall()
    return jsonify([{**dict(r),"data_br":br(r["data"])} for r in rows])

@app.route("/api/cultos", methods=["POST"])
@login_required
def criar_culto():
    # Voluntário pode criar culto (preencher relatório)
    d   = request.get_json(force=True) or {}
    dc  = d.get("data", date.today().isoformat())
    hc  = d.get("hora", datetime.now().strftime("%H:%M"))
    resp= d.get("responsavel","").strip()
    tc  = d.get("tipo_culto","Culto Regular")
    tc_outro = d.get("tipo_outro","").strip()  # descrição quando tipo="Outro"
    if not resp: return jsonify({"erro":"Responsável obrigatório"}),400
    if tc not in TIPOS_CULTO: tc = "Culto Regular"
    # Se "Outro" com descrição, salva como "Outro: descrição"
    if tc == "Outro" and tc_outro:
        tc = f"Outro: {tc_outro}"
    with get_db() as conn:
        cur = conn.execute(
            qmark("""INSERT INTO cultos(data,hora,dia_semana,periodo,tipo_culto,responsavel,
               presentes,visitantes,criancas,observacoes,usuario_id)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)"""),
            (dc,hc,dia_pt(dc),d.get("periodo","Noite"),tc,resp,
             int(d.get("presentes",0)),int(d.get("visitantes",0)),int(d.get("criancas",0)),
             d.get("observacoes",""),session["usuario_id"])
        )
        cid = cur.lastrowid
        for item in conn.execute(
            "SELECT * FROM itens_checklist_padrao ORDER BY categoria,ordem"
        ).fetchall():
            conn.execute(
                qmark("INSERT INTO checklists(culto_id,categoria,item_key,item_descricao,concluido,responsavel) VALUES(?,?,?,?,0,?)"),
                (cid,item["categoria"],item["item_key"],item["descricao"],resp)
            )
        conn.commit()
    return jsonify({"ok":True,"id":cid,"dia_semana":dia_pt(dc)})

@app.route("/api/cultos/<int:cid>", methods=["GET"])
@login_required
def obter_culto(cid):
    with get_db() as conn:
        c = conn.execute(qmark("SELECT id,data,hora,dia_semana,periodo,COALESCE(tipo_culto,'Culto Regular') as tipo_culto,responsavel,presentes,visitantes,criancas,observacoes,criado_em,editado_em,editado_por FROM cultos WHERE id=?"), (cid,)).fetchone()
        if not c: return jsonify({"erro":"Culto não encontrado"}),404
        chks= conn.execute(qmark("SELECT * FROM checklists WHERE culto_id=? ORDER BY categoria,id"), (cid,)).fetchall()
        vis = conn.execute(qmark("SELECT * FROM visitantes WHERE culto_id=? ORDER BY id"), (cid,)).fetchall()
        hist= conn.execute(qmark("SELECT * FROM cultos_historico WHERE culto_id=? ORDER BY alterado_em DESC LIMIT 20"), (cid,)).fetchall()
    row = dict(c); row["data_br"] = br(row["data"])
    return jsonify({
        "culto":      row,
        "checklists": [dict(x) for x in chks],
        "visitantes": [dict(v) for v in vis],
        "historico":  [dict(h) for h in hist]
    })

@app.route("/api/cultos/<int:cid>", methods=["PUT"])
@login_required
def atualizar_culto(cid):
    # Apenas líder e admin podem editar cultos existentes
    if not can_edit():
        return jsonify({"erro":"Apenas líderes e admins podem editar relatórios"}),403
    d = request.get_json(force=True) or {}
    with get_db() as conn:
        antigo = conn.execute(qmark("SELECT * FROM cultos WHERE id=?"), (cid,)).fetchone()
        if not antigo: return jsonify({"erro":"Culto não encontrado"}),404

        # Registra histórico de alterações
        campos = ["presentes","visitantes","criancas","observacoes","periodo","tipo_culto","responsavel"]
        for campo in campos:
            novo_val = str(d.get(campo, antigo[campo] or ""))
            vel      = str(antigo[campo] or "")
            if novo_val != vel:
                conn.execute(
                    qmark("INSERT INTO cultos_historico(culto_id,campo,valor_antes,valor_depois,alterado_por) VALUES(?,?,?,?,?)"),
                    (cid, campo, vel, novo_val, session.get("usuario_nome","?"))
                )

        tc = d.get("tipo_culto", antigo["tipo_culto"])
        tc_outro = d.get("tipo_outro","").strip()
        if tc not in TIPOS_CULTO:
            # Pode ser "Outro: descrição" salvo antes — mantém
            if not str(tc).startswith("Outro:"):
                tc = antigo["tipo_culto"]
        if tc == "Outro" and tc_outro:
            tc = f"Outro: {tc_outro}"

        conn.execute(
            qmark("""UPDATE cultos SET presentes=?,visitantes=?,criancas=?,observacoes=?,
               periodo=?,tipo_culto=?,responsavel=?,editado_em=?,editado_por=?
               WHERE id=?"""),
            (int(d.get("presentes",  antigo["presentes"])),
             int(d.get("visitantes", antigo["visitantes"])),
             int(d.get("criancas",   antigo["criancas"])),
             d.get("observacoes",    antigo["observacoes"] or ""),
             d.get("periodo",        antigo["periodo"]),
             tc,
             d.get("responsavel",    antigo["responsavel"]),
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
             session.get("usuario_nome","?"),
             cid)
        )
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/cultos/<int:cid>", methods=["DELETE"])
@role_required("admin","lider")
def deletar_culto(cid):
    with get_db() as conn:
        conn.execute(qmark("DELETE FROM cultos WHERE id=?"), (cid,)); conn.commit()
    return jsonify({"ok":True})

@app.route("/api/qrcode_fixo", methods=["GET"])
@login_required
def qrcode_fixo():
    """QR Code único permanente — serve para todos os cultos"""
    url = f"{get_base()}/formulario"
    qr  = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H,
                         box_size=8, border=4)
    qr.add_data(url); qr.make(fit=True)
    img = qr.make_image(fill_color="#0A2463", back_color="white")
    buf = io.BytesIO(); img.save(buf,"PNG"); buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode()
    return jsonify({"qrcode":f"data:image/png;base64,{b64}","url":url})

@app.route("/api/cultos/<int:cid>/qrcode", methods=["GET"])
@login_required
def qrcode_culto(cid):
    return qrcode_fixo()

# ═══════════════════════════════════════════════════════════════
# CHECKLIST
# ═══════════════════════════════════════════════════════════════
@app.route("/api/cultos/<int:cid>/checklist", methods=["GET"])
@login_required
def get_checklist(cid):
    with get_db() as conn:
        rows = conn.execute(
            qmark("SELECT * FROM checklists WHERE culto_id=? ORDER BY categoria,id"),(cid,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/checklist/<int:iid>", methods=["PUT"])
@login_required
def atualizar_check(iid):
    d = request.get_json(force=True) or {}
    concluido = 1 if d.get("concluido") else 0
    with get_db() as conn:
        conn.execute(qmark("UPDATE checklists SET concluido=? WHERE id=?"), (concluido,iid))
        conn.commit()
    return jsonify({"ok":True})

# ═══════════════════════════════════════════════════════════════
# VISITANTES
# ═══════════════════════════════════════════════════════════════
@app.route("/api/visitantes", methods=["POST"])
def criar_visitante():
    """Pública — usada pelo QR Code e por usuários logados"""
    d        = request.get_json(force=True) or {}
    nome     = d.get("nome","").strip()
    telefone = d.get("telefone","").strip()
    if not nome or not telefone:
        return jsonify({"erro":"Nome e telefone são obrigatórios"}),400
    with get_db() as conn:
        cur = conn.execute(
            """INSERT INTO visitantes(culto_id,nome,idade,telefone,endereco,
               endereco_padronizado,cidade,bairro,cep,lat,lng,
               como_conheceu,pedido_oracao,quer_visita,data_visita,hora_visita,observacao,origem)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d.get("culto_id") or None, nome, d.get("idade",""), telefone,
             d.get("endereco",""), d.get("endereco_padronizado",""),
             d.get("cidade",""), d.get("bairro",""), d.get("cep",""),
             d.get("lat"), d.get("lng"),
             d.get("como_conheceu",""), d.get("pedido_oracao",""),
             1 if d.get("quer_visita") else 0,
             d.get("data_visita",""), d.get("hora_visita",""),
             d.get("observacao",""), d.get("origem","manual"))
        )
        conn.commit(); vid = cur.lastrowid
    return jsonify({"ok":True,"id":vid})

@app.route("/api/visitantes", methods=["GET"])
@login_required
def listar_visitantes():
    with get_db() as conn:
        rows = conn.execute(
            """SELECT v.*,c.data as culto_data,c.periodo as culto_periodo,c.tipo_culto
               FROM visitantes v LEFT JOIN cultos c ON c.id=v.culto_id
               ORDER BY v.criado_em DESC"""
        ).fetchall()
    return jsonify([{**dict(r),"culto_data_br":br(r["culto_data"]) if r["culto_data"] else ""} for r in rows])

@app.route("/api/visitantes/<int:vid>", methods=["DELETE"])
@role_required("admin","lider")
def deletar_visitante(vid):
    with get_db() as conn:
        conn.execute(qmark("DELETE FROM visitantes WHERE id=?"), (vid,)); conn.commit()
    return jsonify({"ok":True})

@app.route("/api/visitantes/<int:vid>/link", methods=["GET"])
@login_required
def link_visitante(vid):
    with get_db() as conn:
        v = conn.execute(qmark("SELECT * FROM visitantes WHERE id=?"), (vid,)).fetchone()
    if not v: return jsonify({"erro":"Não encontrado"}),404
    v = dict(v)
    q = f"{v.get('endereco','')} {v.get('bairro','')} {v.get('cidade','')}".strip()
    maps = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(q)}"
    tel  = re.sub(r'\D','',v["telefone"])
    if not tel.startswith("55"): tel = "55"+tel
    msg = f"Olá {v['nome']}, tudo bem? Somos da Igreja ABA e ficamos muito felizes com sua visita! 😊"
    wa  = f"https://wa.me/{tel}?text={urllib.parse.quote(msg)}"
    return jsonify({"nome":v["nome"],"telefone":v["telefone"],"maps_link":maps,"whatsapp_link":wa})

# ═══════════════════════════════════════════════════════════════
# GEOCODIFICAÇÃO
# ═══════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════
# ESTOQUE — corrigido
# ═══════════════════════════════════════════════════════════════
@app.route("/api/estoque", methods=["GET"])
@login_required
def listar_estoque():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM estoque ORDER BY fixo DESC,categoria,nome").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/estoque", methods=["POST"])
@role_required("admin","lider")
def criar_estoque():
    d    = request.get_json(force=True) or {}
    nome = d.get("nome","").strip()
    if not nome:
        return jsonify({"erro":"Nome do item é obrigatório"}),400
    try:
        with get_db() as conn:
            qtd     = int(d.get("quantidade",0))
            qtd_min = int(d.get("quantidade_minima",0))
            cur = conn.execute(
                qmark("INSERT INTO estoque(nome,categoria,quantidade,quantidade_minima,unidade,descricao,fixo) VALUES(?,?,?,?,?,?,0)"),
                (nome, d.get("categoria","Geral"), qtd, qtd_min,
                 d.get("unidade","unidade"), d.get("descricao",""))
            )
            conn.commit()
        return jsonify({"ok":True,"id":cur.lastrowid})
    except Exception as e:
        if "UNIQUE" in str(e):
            return jsonify({"erro":f"Já existe um item chamado '{nome}'"}),400
        logger.error(f"Erro criar estoque: {e}")
        return jsonify({"erro":"Erro ao salvar. Tente novamente."}),500

@app.route("/api/estoque/<int:iid>", methods=["PUT"])
@role_required("admin","lider")
def update_estoque(iid):
    d = request.get_json(force=True) or {}
    with get_db() as conn:
        item = conn.execute(qmark("SELECT * FROM estoque WHERE id=?"), (iid,)).fetchone()
        if not item: return jsonify({"erro":"Item não encontrado"}),404
        if is_admin():
            conn.execute(
                qmark("""UPDATE estoque SET nome=?,categoria=?,quantidade=?,quantidade_minima=?,
                   unidade=?,descricao=?,atualizado_em=? WHERE id=?"""),
                (d.get("nome",item["nome"]),d.get("categoria",item["categoria"]),
                 int(d.get("quantidade",item["quantidade"])),
                 int(d.get("quantidade_minima",item["quantidade_minima"])),
                 d.get("unidade",item["unidade"]),d.get("descricao",item["descricao"]),datetime.now().strftime('%Y-%m-%d %H:%M:%S'),iid)
            )
        else:
            conn.execute(
                qmark("UPDATE estoque SET quantidade=?,atualizado_em=? WHERE id=?"),
                (int(d.get("quantidade",item["quantidade"])),datetime.now().strftime('%Y-%m-%d %H:%M:%S'),iid)
            )
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/estoque/<int:iid>", methods=["DELETE"])
@role_required("admin")
def del_estoque(iid):
    with get_db() as conn:
        item = conn.execute(qmark("SELECT fixo FROM estoque WHERE id=?"), (iid,)).fetchone()
        if not item: return jsonify({"erro":"Não encontrado"}),404
        if item["fixo"]:
            return jsonify({"erro":"Itens de Santa Ceia não podem ser excluídos"}),403
        conn.execute(qmark("DELETE FROM estoque WHERE id=?"), (iid,)); conn.commit()
    return jsonify({"ok":True})

# ═══════════════════════════════════════════════════════════════
# CONECTA GC — com líder e edição completa
# ═══════════════════════════════════════════════════════════════
@app.route("/api/gcs", methods=["GET"])
@login_required
def listar_gcs():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM grupos_crescimento WHERE ativo=1 ORDER BY setor,nome"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/gcs/publica", methods=["GET"])
def listar_gcs_publica():
    """Rota pública — usada pelo formulário /relatorio-gc (líderes sem login)"""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id,nome,lider,telefone_lider,setor,cor_hex FROM grupos_crescimento WHERE ativo=1 ORDER BY setor,nome"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/departamentos/publica", methods=["GET"])
def listar_departamentos_publica():
    """Rota pública — usada no auto-cadastro de voluntário na tela de login"""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id,nome,icone FROM departamentos WHERE ativo=1 ORDER BY ordem,nome"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/voluntarios/auto_cadastro", methods=["POST"])
def auto_cadastro_voluntario():
    """Auto-cadastro público: cria usuário (voluntario_escala) + voluntário vinculado"""
    d = request.get_json(force=True) or {}
    nome    = d.get("nome","").strip()
    celular = d.get("celular","").strip()
    email   = d.get("email","").strip().lower()
    senha   = d.get("senha","")
    deptos  = d.get("departamentos",[])  # lista de {id,nome}

    if not nome or not celular or not email or not senha:
        return jsonify({"erro":"Preencha todos os campos."}),400
    if len(senha) < 8:
        return jsonify({"erro":"Senha deve ter no mínimo 8 caracteres."}),400
    if not re.search(r'[A-Z]', senha):
        return jsonify({"erro":"Senha deve ter ao menos uma letra maiúscula."}),400
    if not re.search(r'[0-9]', senha):
        return jsonify({"erro":"Senha deve ter ao menos um número."}),400
    if not deptos:
        return jsonify({"erro":"Selecione pelo menos um departamento."}),400

    # Normaliza departamentos
    nomes_deptos = []
    primeiro_id = None
    for dep in deptos:
        if isinstance(dep, dict):
            if dep.get("nome"): nomes_deptos.append(dep["nome"])
            if primeiro_id is None and dep.get("id"): primeiro_id = dep["id"]
    deptos_txt = ", ".join(nomes_deptos)

    try:
        with get_db() as conn:
            # Verifica e-mail duplicado
            existe = conn.execute(qmark("SELECT id FROM usuarios WHERE email=?"), (email,)).fetchone()
            if existe:
                return jsonify({"erro":"Este e-mail já está cadastrado. Faça login."}),400

            # Cria usuário com cargo voluntario_escala
            cur = conn.execute(
                qmark("INSERT INTO usuarios(nome,email,senha_hash,cargo) VALUES(?,?,?,?)"),
                (nome, email, hs(senha), "voluntario_escala")
            )
            uid = cur.lastrowid
            if USE_PG:
                row = conn.execute(qmark("SELECT id FROM usuarios WHERE email=?"), (email,)).fetchone()
                uid = row["id"] if row else uid

            # Cria voluntário vinculado
            conn.execute(
                qmark("INSERT INTO voluntarios(nome,telefone,departamentos,departamento_id,usuario_id) VALUES(?,?,?,?,?)"),
                (nome, celular, deptos_txt, primeiro_id, uid)
            )
            conn.commit()
        return jsonify({"ok":True,"msg":"Cadastro realizado com sucesso!"})
    except Exception as e:
        logger.error(f"Erro auto_cadastro: {e}")
        if "UNIQUE" in str(e) or "duplicate" in str(e).lower():
            return jsonify({"erro":"Este e-mail já está cadastrado."}),400
        return jsonify({"erro":"Erro ao cadastrar. Tente novamente."}),500

@app.route("/api/gcs", methods=["POST"])
@role_required("admin")
def criar_gc():
    d    = request.get_json(force=True) or {}
    nome = d.get("nome","").strip()
    end  = d.get("endereco","").strip()
    if not nome or not end:
        return jsonify({"erro":"Nome e endereço são obrigatórios"}),400
    bairro = d.get("bairro",""); cidade = d.get("cidade","Alvorada")
    # Geocodifica automaticamente
    lat, lng, _ = geocode_smart(f"{end}, {bairro}, {cidade}")
    cor_map = {"Verde":"#22C55E","Laranja":"#F97316","Amarelo":"#EAB308",
               "Vermelho":"#EF4444","Azul":"#3B82F6","Roxo":"#A855F7"}
    setor = d.get("setor","Verde")
    gc_pai = d.get("gc_pai_id") or None
    supervisor = d.get("supervisor","").strip()
    with get_db() as conn:
        cur = conn.execute(
            qmark("INSERT INTO grupos_crescimento(nome,lider,telefone_lider,endereco,bairro,cidade,setor,cor_hex,lat,lng,gc_pai_id,supervisor) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)"),
            (nome, d.get("lider",""), d.get("telefone_lider",""), end, bairro, cidade, setor,
             d.get("cor_hex", cor_map.get(setor,"#22C55E")), lat, lng, gc_pai, supervisor)
        )
        conn.commit()
        new_id = cur.lastrowid
        if USE_PG:
            row = conn.execute(qmark("SELECT id FROM grupos_crescimento WHERE nome=? ORDER BY id DESC"),(nome,)).fetchone()
            new_id = row["id"] if row else new_id
    return jsonify({"ok":True,"id":new_id})

@app.route("/api/gcs/<int:gid>", methods=["PUT"])
@role_required("admin")
def atualizar_gc(gid):
    d = request.get_json(force=True) or {}
    with get_db() as conn:
        gc = conn.execute(qmark("SELECT * FROM grupos_crescimento WHERE id=?"), (gid,)).fetchone()
        if not gc: return jsonify({"erro":"GC não encontrado"}),404
        novo_end = d.get("endereco", gc["endereco"])
        novo_bai = d.get("bairro",   gc["bairro"])
        nova_cid = d.get("cidade",   gc["cidade"])
        lat, lng = gc["lat"], gc["lng"]
        if novo_end != gc.get("endereco","") or novo_bai != gc.get("bairro",""):
            new_lat, new_lng, _ = geocode_smart(f"{novo_end}, {novo_bai}, {nova_cid}")
            if new_lat:
                lat, lng = new_lat, new_lng
        conn.execute(
            qmark("""UPDATE grupos_crescimento SET nome=?,lider=?,telefone_lider=?,endereco=?,bairro=?,cidade=?,
               setor=?,cor_hex=?,lat=?,lng=?,gc_pai_id=?,supervisor=?,ativo=? WHERE id=?"""),
            (d.get("nome",gc["nome"]),
             d.get("lider",gc.get("lider","")),
             d.get("telefone_lider",gc.get("telefone_lider","")),
             novo_end, novo_bai, nova_cid,
             d.get("setor",gc["setor"]), d.get("cor_hex",gc["cor_hex"]),
             lat, lng,
             d.get("gc_pai_id") if d.get("gc_pai_id") else None,
             d.get("supervisor", gc["supervisor"] if "supervisor" in gc.keys() else ""),
             int(d.get("ativo",gc["ativo"])), gid)
        )
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/gcs/<int:gid>", methods=["DELETE"])
@role_required("admin")
def del_gc(gid):
    with get_db() as conn:
        conn.execute(qmark("UPDATE grupos_crescimento SET ativo=0 WHERE id=?"), (gid,)); conn.commit()
    return jsonify({"ok":True})

# ═══════════════════════════════════════════════════════════════
# GESTÃO DE GC: Multiplicação, Redes, Supervisão, Integrantes
# ═══════════════════════════════════════════════════════════════
@app.route("/api/gcs/arvore", methods=["GET"])
@role_required("admin","lider")
def gcs_arvore():
    """Retorna a árvore de multiplicação dos GCs (pai → filhos)"""
    with get_db() as conn:
        gcs = [dict(r) for r in conn.execute(
            "SELECT id,nome,lider,setor,cor_hex,gc_pai_id,supervisor FROM grupos_crescimento WHERE ativo=1 ORDER BY nome"
        ).fetchall()]
    # Monta estrutura de árvore
    by_id = {g["id"]: {**g, "filhos": []} for g in gcs}
    raizes = []
    for g in gcs:
        pai = g.get("gc_pai_id")
        if pai and pai in by_id:
            by_id[pai]["filhos"].append(by_id[g["id"]])
        else:
            raizes.append(by_id[g["id"]])
    return jsonify({"arvore": raizes, "total": len(gcs)})

@app.route("/api/gcs/metricas_redes", methods=["GET"])
@role_required("admin","lider")
def gcs_metricas_redes():
    """Métricas por rede (setor/cor): quantidade, multiplicações, crescimento"""
    with get_db() as conn:
        gcs = [dict(r) for r in conn.execute(
            "SELECT id,nome,setor,cor_hex,gc_pai_id,criado_em FROM grupos_crescimento WHERE ativo=1"
        ).fetchall()]
    redes = {}
    for g in gcs:
        rede = g.get("setor") or "Sem rede"
        if rede not in redes:
            redes[rede] = {"rede": rede, "cor": g.get("cor_hex","#94A3B8"),
                           "total_gcs": 0, "multiplicacoes": 0}
        redes[rede]["total_gcs"] += 1
        if g.get("gc_pai_id"):
            redes[rede]["multiplicacoes"] += 1
    lista = sorted(redes.values(), key=lambda x: x["total_gcs"], reverse=True)
    maior = lista[0]["rede"] if lista else "—"
    return jsonify({
        "redes": lista,
        "total_gcs": len(gcs),
        "total_multiplicacoes": sum(r["multiplicacoes"] for r in lista),
        "rede_maior": maior
    })

@app.route("/api/gcs/por_supervisor", methods=["GET"])
@role_required("admin","lider")
def gcs_por_supervisor():
    """Agrupa GCs por supervisor para relatório"""
    with get_db() as conn:
        gcs = [dict(r) for r in conn.execute(
            "SELECT id,nome,lider,setor,supervisor FROM grupos_crescimento WHERE ativo=1 ORDER BY supervisor,nome"
        ).fetchall()]
    por_sup = {}
    for g in gcs:
        sup = (g.get("supervisor") or "").strip() or "Sem supervisor"
        if sup not in por_sup:
            por_sup[sup] = {"supervisor": sup, "gcs": []}
        por_sup[sup]["gcs"].append({"id":g["id"],"nome":g["nome"],"lider":g.get("lider",""),"setor":g.get("setor","")})
    lista = sorted(por_sup.values(), key=lambda x: (-len(x["gcs"]), x["supervisor"]))
    return jsonify({"supervisores": lista, "total": len(gcs)})

# ── GENEALOGIA COMPLETA DOS GCs ──
def _calcular_genealogia():
    """Carrega todos os GCs e calcula nível, filhos diretos e descendentes totais"""
    with get_db() as conn:
        gcs = [dict(r) for r in conn.execute(
            "SELECT id,nome,lider,setor,cor_hex,gc_pai_id,supervisor,criado_em FROM grupos_crescimento WHERE ativo=1 ORDER BY nome"
        ).fetchall()]
    by_id = {g["id"]: g for g in gcs}
    filhos_map = {}
    for g in gcs:
        pai = g.get("gc_pai_id")
        if pai and pai in by_id:
            filhos_map.setdefault(pai, []).append(g["id"])

    def nivel(gid, visto=None):
        visto = visto or set()
        if gid in visto: return 1
        visto.add(gid)
        pai = by_id.get(gid, {}).get("gc_pai_id")
        if pai and pai in by_id:
            return 1 + nivel(pai, visto)
        return 1

    def contar_descendentes(gid, visto=None):
        visto = visto or set()
        total = 0
        for fid in filhos_map.get(gid, []):
            if fid in visto: continue
            visto.add(fid)
            total += 1 + contar_descendentes(fid, visto)
        return total

    for g in gcs:
        g["nivel"] = nivel(g["id"])
        g["qtd_filhos"] = len(filhos_map.get(g["id"], []))
        g["qtd_descendentes"] = contar_descendentes(g["id"])
        g["gc_pai_nome"] = by_id.get(g.get("gc_pai_id"), {}).get("nome", "") if g.get("gc_pai_id") else ""
    return gcs, by_id, filhos_map

@app.route("/api/gcs/genealogia", methods=["GET"])
@role_required("admin","lider")
def gcs_genealogia():
    """Retorna árvore + cards estratégicos de multiplicação"""
    ini = request.args.get("data_ini","")
    fim = request.args.get("data_fim","")
    gcs, by_id, filhos_map = _calcular_genealogia()

    # Monta árvore aninhada
    def montar_no(gid):
        g = by_id[gid]
        return {
            "id": g["id"], "nome": g["nome"], "lider": g.get("lider",""),
            "setor": g.get("setor",""), "cor_hex": g.get("cor_hex","#94A3B8"),
            "supervisor": g.get("supervisor",""),
            "qtd_filhos": g["qtd_filhos"], "qtd_descendentes": g["qtd_descendentes"],
            "nivel": g["nivel"],
            "filhos": [montar_no(fid) for fid in filhos_map.get(gid, [])]
        }
    raizes = [montar_no(g["id"]) for g in gcs if not (g.get("gc_pai_id") and g["gc_pai_id"] in by_id)]

    # Cards estratégicos
    multiplicados = [g for g in gcs if g["qtd_filhos"] > 0]
    geracoes = max([g["nivel"] for g in gcs], default=0)
    maior_desc = max(gcs, key=lambda x: x["qtd_descendentes"], default=None)

    # Métricas por rede
    redes = {}
    for g in gcs:
        rede = g.get("setor") or "Sem rede"
        if rede not in redes:
            redes[rede] = {"rede": rede, "cor": g.get("cor_hex","#94A3B8"), "total_gcs": 0, "multiplicacoes": 0}
        redes[rede]["total_gcs"] += 1
        if g.get("gc_pai_id"): redes[rede]["multiplicacoes"] += 1
    redes_lista = sorted(redes.values(), key=lambda x: (-x["total_gcs"], x["rede"]))

    # Multiplicações no período (por criado_em do GC filho)
    mult_periodo = 0
    for g in gcs:
        if g.get("gc_pai_id"):
            cdt = str(g.get("criado_em",""))[:10]
            if (not ini or cdt >= ini) and (not fim or cdt <= fim):
                mult_periodo += 1

    cards = {
        "total_ativos": len(gcs),
        "total_multiplicados": len(multiplicados),
        "geracoes": geracoes,
        "maior_descendencia": {"nome": maior_desc["nome"], "qtd": maior_desc["qtd_descendentes"]} if maior_desc else {"nome":"—","qtd":0},
        "rede_maior": redes_lista[0]["rede"] if redes_lista else "—",
        "multiplicacoes_periodo": mult_periodo
    }
    # Linha do tempo das multiplicações (GCs com pai), ordenada por data de criação
    MESES_ABR = ["","Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    timeline = []
    for g in gcs:
        if g.get("gc_pai_id") and g["gc_pai_id"] in by_id:
            cdt = str(g.get("criado_em",""))[:10]
            # Rótulo amigável "Mês Ano"
            label = cdt
            try:
                dd = datetime.strptime(cdt, "%Y-%m-%d")
                label = f"{MESES_ABR[dd.month]} {dd.year}"
            except Exception:
                pass
            timeline.append({
                "data": cdt,
                "data_label": label,
                "nome": g["nome"],
                "rede": g.get("setor",""),
                "cor": g.get("cor_hex","#94A3B8"),
                "pai": by_id[g["gc_pai_id"]]["nome"]
            })
    timeline.sort(key=lambda x: x["data"])

    return jsonify({"arvore": raizes, "cards": cards, "redes": redes_lista, "timeline": timeline})

@app.route("/api/gcs/genealogia/excel", methods=["GET"])
@role_required("admin","lider")
def gcs_genealogia_excel():
    """Exporta genealogia dos GCs em Excel com colunas hierárquicas"""
    gcs, by_id, filhos_map = _calcular_genealogia()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Genealogia GCs"
    azul = PatternFill("solid",fgColor="0A2463"); cinza = PatternFill("solid",fgColor="EBF5FF")
    branco = PatternFill("solid",fgColor="FFFFFF")
    tw = Font(color="FFFFFF",bold=True,size=11); tn = Font(size=10); tb = Font(size=10,bold=True)
    centro = Alignment(horizontal="center",vertical="center"); esq = Alignment(horizontal="left",vertical="center")
    borda = Border(*[Side(style="thin",color="D1D5DB")]*4)
    cols = ["GC","Rede","Supervisor","GC Origem","Nível Hierárquico","Qtd. Filhos","Qtd. Descendentes"]
    larg = [26,14,20,24,16,12,16]
    for i,(c,l) in enumerate(zip(cols,larg),1):
        ws.column_dimensions[get_column_letter(i)].width = l
        cell = ws.cell(row=1,column=i,value=c); cell.fill=azul; cell.font=tw; cell.alignment=centro; cell.border=borda
    ws.row_dimensions[1].height = 26
    gcs_ord = sorted(gcs, key=lambda x: (x["nivel"], x["nome"]))
    for i,g in enumerate(gcs_ord,2):
        fill = cinza if i%2==0 else branco
        vals = [g["nome"], g.get("setor",""), g.get("supervisor","") or "—",
                g.get("gc_pai_nome","") or "— (raiz)", g["nivel"], g["qtd_filhos"], g["qtd_descendentes"]]
        for j,v in enumerate(vals,1):
            c = ws.cell(row=i,column=j,value=v)
            c.fill=fill; c.font=tb if j==1 else tn
            c.alignment=esq if j in (1,2,3,4) else centro; c.border=borda
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"genealogia_gcs_{date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/gcs/genealogia/pdf", methods=["GET"])
@role_required("admin","lider")
def gcs_genealogia_pdf():
    """Exporta genealogia + árvore visual em PDF (HTML para impressão)"""
    gcs, by_id, filhos_map = _calcular_genealogia()

    # Árvore em HTML recursiva
    def render_no(gid, prof=0):
        g = by_id[gid]
        cor = g.get("cor_hex","#94A3B8")
        badge = f'<span style="background:{cor};color:#fff;padding:1px 7px;border-radius:20px;font-size:9px;font-weight:700">{g.get("setor","")}</span>'
        desc = f' · {g["qtd_descendentes"]} descendente(s)' if g["qtd_descendentes"] else ""
        filhos = filhos_map.get(gid, [])
        html = f'''<div class="node" style="margin-left:{prof*22}px">
          <div class="node-box" style="border-left:4px solid {cor}">
            <strong>{g["nome"]}</strong> {badge}
            <div class="node-meta">👤 {g.get("lider","—")} · Nível {g["nivel"]}{desc}{(" · 👁️ "+g.get("supervisor","")) if g.get("supervisor") else ""}</div>
          </div>'''
        for fid in filhos:
            html += render_no(fid, prof+1)
        html += '</div>'
        return html

    raizes = [g["id"] for g in gcs if not (g.get("gc_pai_id") and g["gc_pai_id"] in by_id)]
    arvore_html = "".join(render_no(r) for r in raizes) or '<p style="text-align:center;color:#94A3B8;padding:30px">Nenhum GC cadastrado.</p>'

    multiplicados = len([g for g in gcs if g["qtd_filhos"]>0])
    geracoes = max([g["nivel"] for g in gcs], default=0)

    html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Genealogia dos GCs</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;background:#F1F5F9;color:#1E293B}}
.btn-print{{display:block;width:calc(100% - 32px);margin:16px auto 0;padding:14px;background:#0A2463;color:#fff;border:none;border-radius:12px;font-size:15px;font-weight:700;cursor:pointer}}
.hint{{text-align:center;font-size:11px;color:#94A3B8;margin:6px 16px 14px}}
.header{{background:linear-gradient(135deg,#0A2463,#1B4FA8);color:#fff;padding:22px 16px;margin-bottom:16px}}
.header h1{{font-size:20px;font-weight:900;letter-spacing:1px}}
.header .meta{{font-size:12px;opacity:.7;margin-top:4px}}
.stats{{display:flex;gap:10px;padding:0 16px 16px;flex-wrap:wrap}}
.stat{{flex:1;min-width:110px;background:#fff;border-radius:12px;padding:12px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.stat .v{{font-size:24px;font-weight:800;color:#0A2463}} .stat .l{{font-size:10px;color:#64748B;text-transform:uppercase;margin-top:3px}}
.arvore{{background:#fff;border-radius:12px;margin:0 16px 16px;padding:16px}}
.arvore-titulo{{font-size:14px;font-weight:800;color:#0A2463;margin-bottom:14px}}
.node-box{{background:#F8FAFF;border-radius:8px;padding:8px 12px;margin-bottom:6px}}
.node-box strong{{font-size:13px;color:#0F2747}}
.node-meta{{font-size:11px;color:#64748B;margin-top:2px}}
.footer{{text-align:center;font-size:10px;color:#94A3B8;padding:16px}}
@media print{{.btn-print,.hint{{display:none}}body{{background:#fff}}}}
</style></head><body>
<button class="btn-print" onclick="window.print()">💾 Salvar como PDF</button>
<p class="hint">Toque nos 3 pontos → Imprimir → Salvar como PDF</p>
<div class="header"><h1>GENEALOGIA DOS GCs</h1>
  <div class="meta">Igreja ABA · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}</div></div>
<div class="stats">
  <div class="stat"><div class="v">{len(gcs)}</div><div class="l">GCs Ativos</div></div>
  <div class="stat"><div class="v">{multiplicados}</div><div class="l">Multiplicados</div></div>
  <div class="stat"><div class="v">{geracoes}</div><div class="l">Gerações</div></div>
</div>
<div class="arvore">
  <div class="arvore-titulo">🌳 Árvore de Multiplicação</div>
  {arvore_html}
</div>
<div class="footer">Igreja ABA — Um Lar Para Pertencer</div>
</body></html>"""
    resp = make_response(html); resp.headers["Content-Type"]="text/html; charset=utf-8"
    return resp

# ── INTEGRANTES DO GC ──
def _capitalizar_nome(nome):
    """Padroniza nome: primeira letra de cada palavra em maiúsculo"""
    palavras = nome.strip().split()
    minusculas = {"de","da","do","das","dos","e"}
    out = []
    for i, p in enumerate(palavras):
        pl = p.lower()
        if i > 0 and pl in minusculas:
            out.append(pl)
        else:
            out.append(pl.capitalize())
    return " ".join(out)

@app.route("/api/gcs/<int:gid>/integrantes", methods=["GET"])
@role_required("admin","lider")
def listar_integrantes(gid):
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(
            qmark("SELECT id,nome,telefone FROM gc_integrantes WHERE gc_id=? AND ativo=1 ORDER BY nome"),
            (gid,)
        ).fetchall()]
    return jsonify(rows)

@app.route("/api/gcs/<int:gid>/integrantes", methods=["POST"])
@role_required("admin","lider")
def add_integrante(gid):
    d = request.get_json(force=True) or {}
    nome = d.get("nome","").strip()
    # Validação: exige nome + sobrenome
    if len(nome.split()) < 2:
        return jsonify({"erro":"Informe nome e sobrenome completos (ex: Pedro Pedreira)."}),400
    nome = _capitalizar_nome(nome)
    with get_db() as conn:
        conn.execute(
            qmark("INSERT INTO gc_integrantes(gc_id,nome,telefone) VALUES(?,?,?)"),
            (gid, nome, d.get("telefone","").strip())
        )
        conn.commit()
    return jsonify({"ok":True,"nome":nome})

@app.route("/api/gcs/integrantes/<int:iid>", methods=["DELETE"])
@role_required("admin","lider")
def del_integrante(iid):
    with get_db() as conn:
        conn.execute(qmark("UPDATE gc_integrantes SET ativo=0 WHERE id=?"), (iid,))
        conn.commit()
    return jsonify({"ok":True})



# ── Coordenadas dos bairros de Alvorada/RS (offline, sem API) ─────────────
_BAIRROS = {
    "jardim algarve":(-30.0240,-51.0808),"algarve":(-30.0240,-51.0808),
    "porto verde":(-30.0330,-51.0760),"intersul":(-30.0195,-51.0720),
    "jardim porto alegre":(-30.0280,-51.0742),"centro":(-30.0220,-51.0780),
    "grajaú":(-30.0180,-51.0760),"grajau":(-30.0180,-51.0760),
    "jardim presidente":(-30.0260,-51.0820),"bom princípio":(-30.0300,-51.0790),
    "bom principio":(-30.0300,-51.0790),"são paulo":(-30.0250,-51.0835),
    "sao paulo":(-30.0250,-51.0835),"niterói":(-30.0210,-51.0800),
    "niteroi":(-30.0210,-51.0800),"nova alvorada":(-30.0350,-51.0800),
    "parque amador":(-30.0370,-51.0780),"santa fé":(-30.0290,-51.0760),
    "santa fe":(-30.0290,-51.0760),"morada do vale":(-30.0310,-51.0820),
    "sete de setembro":(-30.0200,-51.0740),"universitário":(-30.0270,-51.0850),
    "universitario":(-30.0270,-51.0850),"residencial":(-30.0320,-51.0790),
    "são luís":(-30.0265,-51.0770),"sao luis":(-30.0265,-51.0770),
    "jardim nova alvorada":(-30.0360,-51.0810),"nova":(-30.0350,-51.0800),
}

def geocode_por_bairro(texto):
    """Retorna (lat, lng, bairro_nome) baseado no bairro detectado no texto."""
    t = texto.lower().strip()
    # Remove acentos para matching mais tolerante
    import unicodedata
    t_norm = ''.join(c for c in unicodedata.normalize('NFD',t) if unicodedata.category(c) != 'Mn')
    for bairro, coords in _BAIRROS.items():
        b_norm = ''.join(c for c in unicodedata.normalize('NFD',bairro) if unicodedata.category(c) != 'Mn')
        if b_norm in t_norm:
            return coords[0], coords[1], bairro.title()
    # Tenta por palavras
    words = t_norm.replace(","," ").replace("-"," ").split()
    for bairro, coords in _BAIRROS.items():
        b_norm = ''.join(c for c in unicodedata.normalize('NFD',bairro) if unicodedata.category(c) != 'Mn')
        for w in words:
            if len(w) >= 5 and w in b_norm:
                return coords[0], coords[1], bairro.title()
    return None, None, ""

@app.route("/api/gcs/calcular_proximo", methods=["POST"])
@login_required
def calcular_gc():
    d   = request.get_json(force=True) or {}
    q   = d.get("query","").strip()  # campo único de busca
    end = d.get("endereco","").strip()
    bai = d.get("bairro","").strip()
    cid = d.get("cidade","Alvorada").strip()
    if not q and not end:
        return jsonify({"erro":"Digite um endereço"}),400
    busca = q or f"{end}, {bai}, {cid}"
    # 1. Tenta geocode externo (Nominatim)
    lat_v, lng_v, display_v = geocode_smart(busca, "Alvorada")

    # 2. Se falhar, usa bairro detectado no texto (offline, sempre funciona)
    if not lat_v:
        lat_v, lng_v, bairro_det = geocode_por_bairro(busca)
        if lat_v:
            display_v = bairro_det
            logger.info(f"Usando bairro detectado: {bairro_det}")

    # 3. Se ainda falhar, retorna erro claro
    if not lat_v:
        return jsonify({
            "erro": "Bairro não reconhecido.",
            "dica": "Digite o nome do bairro. Ex: Jardim Algarve, Porto Verde, Intersul, Centro"
        }), 422
    with get_db() as conn:
        gcs = conn.execute(
            "SELECT * FROM grupos_crescimento WHERE ativo=1 AND lat IS NOT NULL AND lat!=0"
        ).fetchall()
    if not gcs:
        return jsonify({"erro":"GCs ainda sem coordenadas. Clique em 'Geocodificar GCs'."}),422
    results = []
    for gc in gcs:
        dist = haversine(lat_v,lng_v,gc["lat"],gc["lng"])
        _o = urllib.parse.quote(display_v or busca)
        gc_end = gc['endereco'] + ", " + gc['bairro'] + ", " + gc['cidade'] + ", RS"
        _d = urllib.parse.quote(gc_end)
        rota = "https://www.google.com/maps/dir/?api=1&origin=" + _o + "&destination=" + _d + "&travelmode=driving"
        gc_nome_enc = urllib.parse.quote(gc['nome'])
        msg_wa = "Rota para " + gc['nome'] + "%0A" + "Endereco: " + gc_end + "%0ADistancia: " + str(round(dist,2)) + " km%0A%0AMaps: " + rota
        wa_rota = "https://wa.me/?text=" + msg_wa
        results.append({**dict(gc),"distancia_km":round(dist,2),"rota_link":rota,"wa_rota":wa_rota})
    results.sort(key=lambda x:x["distancia_km"])
    return jsonify({"ok":True,
                    "visitante":{"lat":lat_v,"lng":lng_v,"endereco":display_v or busca},
                    "gcs":results,"mais_proximo":results[0]})

@app.route("/api/gcs/direcionar", methods=["POST"])
@login_required
def direcionar():
    d = request.get_json(force=True) or {}
    gc_id = d.get("gc_id")
    whatsapp_lider = None
    with get_db() as conn:
        conn.execute(
            qmark("INSERT INTO gc_direcionamentos(visitante_id,gc_id,visitante_nome,gc_nome,distancia_km) VALUES(?,?,?,?,?)"),
            (d.get("visitante_id"),gc_id,d.get("visitante_nome",""),d.get("gc_nome",""),d.get("distancia_km"))
        )
        if d.get("visitante_id"):
            conn.execute(qmark("UPDATE visitantes SET observacao=? WHERE id=?"),
                         (f"Direcionado para: {d.get('gc_nome','')}", d.get("visitante_id")))
        # Busca telefone do líder para gerar link WhatsApp
        if gc_id:
            gc_row = conn.execute(qmark("SELECT lider,telefone_lider FROM grupos_crescimento WHERE id=?"), (gc_id,)).fetchone()
            if gc_row and gc_row.get("telefone_lider","").strip():
                tel = re.sub(r"\D","",gc_row["telefone_lider"])
                if not tel.startswith("55"): tel = "55" + tel
                vis_nome = d.get("visitante_nome","Visitante")
                gc_nome  = d.get("gc_nome","")
                dist     = d.get("distancia_km","")
                rota     = d.get("rota_link","")
                lider_nome = gc_row.get("lider","Lider") or "Lider"
                msg = (
                    "Ola " + lider_nome + "!\n"
                    "Temos um visitante para o " + gc_nome + ".\n\n"
                    "Nome do visitante: " + vis_nome + "\n"
                    "Distancia: " + str(dist) + " km\n"
                    "Rota no Maps: " + rota + "\n\n"
                    "Enviado pelo sistema Igreja ABA"
                )
                whatsapp_lider = f"https://wa.me/{tel}?text={urllib.parse.quote(msg)}"
        conn.commit()
    return jsonify({"ok":True, "whatsapp_lider": whatsapp_lider})

@app.route("/api/gcs/direcionamentos", methods=["GET"])
@login_required
def direcionamentos():
    with get_db() as conn:
        rows = conn.execute(
            """SELECT d.*,gc.cor_hex FROM gc_direcionamentos d
               LEFT JOIN grupos_crescimento gc ON gc.id=d.gc_id
               ORDER BY d.criado_em DESC LIMIT 100"""
        ).fetchall()
    return jsonify([dict(r) for r in rows])

# ═══════════════════════════════════════════════════════════════
# DASHBOARD — Analytics
# ═══════════════════════════════════════════════════════════════
@app.route("/api/dashboard", methods=["GET"])
@role_required("lider","admin")
def dashboard():
    ano  = request.args.get("ano",  str(date.today().year))
    mes  = request.args.get("mes",  "")
    tpc  = request.args.get("tipo", "")
    per  = request.args.get("periodo", "")
    ini  = request.args.get("data_ini", "")
    fim  = request.args.get("data_fim", "")

    # Monta cláusula WHERE comum para resumo/por_tipo
    where = " WHERE 1=1"
    wparams = []
    if per: where += " AND periodo=?"; wparams.append(per)
    if ini: where += " AND data>=?"; wparams.append(ini)
    if fim: where += " AND data<=?"; wparams.append(fim)
    if tpc: where += " AND tipo_culto=?"; wparams.append(tpc)

    with get_db() as conn:
        # Resumo geral (com filtros)
        try:
            resumo = dict(conn.execute(qmark(f"""SELECT COUNT(*) as total_cultos,
               COALESCE(SUM(presentes),0) as total_presentes,
               COALESCE(SUM(visitantes),0) as total_visitantes,
               COALESCE(SUM(criancas),0) as total_criancas,
               COALESCE(ROUND(CAST(AVG(presentes) AS {'NUMERIC' if USE_PG else 'REAL'}),1),0) as media_presentes,
               COALESCE(ROUND(CAST(AVG(visitantes) AS {'NUMERIC' if USE_PG else 'REAL'}),1),0) as media_visitantes,
               COALESCE(ROUND(CAST(AVG(criancas) AS {'NUMERIC' if USE_PG else 'REAL'}),1),0) as media_criancas
               FROM cultos{where}"""), wparams).fetchone() or {})
        except Exception as e:
            logger.warning(f"Dashboard resumo erro: {e}")
            resumo = {}

        # Evolução mensal — compatível PG e SQLite
        if USE_PG:
            sql_m = """SELECT to_char(data::date,'YYYY-MM') as mes,
                       COALESCE(SUM(presentes),0) as presentes,
                       COALESCE(SUM(visitantes),0) as visitantes,
                       COALESCE(SUM(criancas),0) as criancas,
                       COUNT(*) as cultos
                       FROM cultos WHERE to_char(data::date,'YYYY')=%s
                       GROUP BY mes ORDER BY mes"""
        else:
            sql_m = """SELECT strftime('%Y-%m',data) as mes,
                       SUM(presentes) as presentes, SUM(visitantes) as visitantes,
                       SUM(criancas) as criancas, COUNT(*) as cultos
                       FROM cultos WHERE strftime('%Y',data)=?
                       GROUP BY mes ORDER BY mes"""
        try:
            mensal = [dict(r) for r in conn.execute(sql_m,(ano,)).fetchall()]
        except Exception as e:
            logger.warning(f"Dashboard mensal erro: {e}")
            mensal = []

        # Por tipo de culto (com filtros)
        try:
            por_tipo = [dict(r) for r in conn.execute(qmark(
                f"""SELECT COALESCE(tipo_culto,'Culto Regular') as tipo_culto,
                   COUNT(*) as qtd,
                   COALESCE(SUM(presentes),0) as total_presentes,
                   COALESCE(SUM(visitantes),0) as total_visitantes
                   FROM cultos{where} GROUP BY tipo_culto ORDER BY total_presentes DESC"""
            ), wparams).fetchall()]
            # Adiciona media_presentes manualmente
            for t in por_tipo:
                t['media_presentes'] = round(t['total_presentes']/max(t['qtd'],1),1)
        except Exception as e:
            logger.warning(f"Dashboard por_tipo erro: {e}")
            por_tipo = []

        # Top GCs
        try:
            top_gcs = [dict(r) for r in conn.execute(
                """SELECT gc_nome, COUNT(*) as direcionamentos
                   FROM gc_direcionamentos GROUP BY gc_nome
                   ORDER BY direcionamentos DESC LIMIT 5"""
            ).fetchall()]
        except:
            top_gcs = []

        # Últimos 5 cultos
        try:
            ultimos = [dict(r) for r in conn.execute(
                qmark("SELECT id,data,hora,dia_semana,periodo,COALESCE(tipo_culto,'Culto Regular') as tipo_culto,responsavel,presentes,visitantes,criancas,observacoes,criado_em FROM cultos ORDER BY data DESC,hora DESC LIMIT 5")
            ).fetchall()]
        except Exception as e:
            logger.warning(f"Dashboard ultimos erro: {e}")
            ultimos = []

        # Crescimento mês a mês
        try:
            if USE_PG:
                meses_list = [dict(r) for r in conn.execute(
                    """SELECT to_char(data::date,'YYYY-MM') as mes,
                       COALESCE(SUM(presentes),0) as presentes
                       FROM cultos GROUP BY mes ORDER BY mes DESC LIMIT 12"""
                ).fetchall()]
            else:
                meses_list = [dict(r) for r in conn.execute(
                    """SELECT strftime('%Y-%m',data) as mes, SUM(presentes) as presentes
                       FROM cultos GROUP BY mes ORDER BY mes DESC LIMIT 12"""
                ).fetchall()]
        except:
            meses_list = []

        # Visitantes por mês
        try:
            if USE_PG:
                vis_mensal = [dict(r) for r in conn.execute(
                    """SELECT to_char(data::date,'MM') as mes,
                       COALESCE(SUM(visitantes),0) as visitantes
                       FROM cultos WHERE to_char(data::date,'YYYY')=%s
                       GROUP BY mes ORDER BY mes""", (ano,)
                ).fetchall()]
            else:
                vis_mensal = [dict(r) for r in conn.execute(
                    """SELECT strftime('%m',data) as mes, SUM(visitantes) as visitantes
                       FROM cultos WHERE strftime('%Y',data)=? GROUP BY mes ORDER BY mes""", (ano,)
                ).fetchall()]
        except:
            vis_mensal = []

    # Insights automáticos
    insights = []
    if mensal:
        melhor = max(mensal, key=lambda x: x["presentes"])
        insights.append(f"🏆 {_mes_nome(melhor['mes'])} foi o mês com maior presença ({melhor['presentes']} pessoas).")
    if por_tipo:
        t = por_tipo[0]
        insights.append(f"📌 '{t['tipo_culto']}' é o tipo de culto com mais presença (média: {t['media_presentes']}).")
    if top_gcs:
        insights.append(f"🌟 {top_gcs[0]['gc_nome']} recebeu mais visitantes direcionados ({top_gcs[0]['direcionamentos']}).")
    if len(meses_list) >= 2:
        atual = meses_list[0]["presentes"] or 0
        ant   = meses_list[1]["presentes"] or 1
        diff  = round((atual - ant) / ant * 100, 1) if ant else 0
        sinal = "cresceu" if diff >= 0 else "caiu"
        insights.append(f"📈 A presença {sinal} {abs(diff)}% em relação ao mês anterior.")

    # Converte Decimal para int/float para JSON
    def safe(v):
        if v is None: return 0
        try: return float(v) if '.' in str(v) else int(v)
        except: return 0

    resumo_clean = {k: safe(v) for k,v in resumo.items()}
    mensal_clean = [{k: safe(v) if k != "mes" else v for k,v in m.items()} for m in mensal]
    por_tipo_clean = [{k: safe(v) if k != "tipo_culto" else v for k,v in t.items()} for t in por_tipo]

    return jsonify({
        "resumo":     resumo_clean,
        "mensal":     mensal_clean,
        "por_tipo":   por_tipo_clean,
        "top_gcs":    top_gcs,
        "ultimos":    [{**u,"data_br":br(u["data"])} for u in ultimos],
        "vis_mensal": vis_mensal,
        "insights":   insights
    })

def _mes_nome(ym):
    meses = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    try:
        m = int(ym.split("-")[1]) - 1
        return f"{meses[m]}/{ym.split('-')[0]}"
    except: return ym

# ═══════════════════════════════════════════════════════════════
# RESUMO
# ═══════════════════════════════════════════════════════════════
@app.route("/api/resumo", methods=["GET"])
@login_required
def resumo():
    with get_db() as conn:
        if USE_PG:
            r = conn.execute("""SELECT
               COUNT(*) as total_cultos,
               COALESCE(SUM(presentes),0) as total_presentes,
               COALESCE(SUM(visitantes),0) as total_visitantes,
               COALESCE(SUM(criancas),0) as total_criancas,
               COALESCE(ROUND(CAST(AVG(presentes) AS NUMERIC),1),0) as media_presentes,
               COALESCE(ROUND(CAST(AVG(visitantes) AS NUMERIC),1),0) as media_visitantes,
               COALESCE(ROUND(CAST(AVG(criancas) AS NUMERIC),1),0) as media_criancas
               FROM cultos""").fetchone()
        else:
            r = conn.execute("SELECT * FROM v_resumo_geral").fetchone()
        ult = conn.execute("SELECT id,data,hora,dia_semana,periodo,COALESCE(tipo_culto,'Culto Regular') as tipo_culto,responsavel,presentes,visitantes,criancas,observacoes,criado_em,editado_em,editado_por FROM cultos ORDER BY data DESC,hora DESC LIMIT 5").fetchall()
        pp  = conn.execute("""SELECT periodo,COUNT(*) as qtd,
                            COALESCE(ROUND(CAST(AVG(presentes) AS NUMERIC),1),0) as mp,
                            COALESCE(SUM(presentes),0) as tp FROM cultos GROUP BY periodo""").fetchall()
    # Converte Decimal/None para tipos JSON-serializáveis
    def safe(v):
        if v is None: return 0
        try: return float(v) if '.' in str(v) else int(v)
        except: return 0
    geral = {}
    if r:
        rd = dict(r)
        geral = {k: safe(v) for k,v in rd.items()}
    return jsonify({
        "geral":      geral,
        "ultimos":    [{**dict(u),"data_br":br(u["data"])} for u in ult],
        "por_periodo":[{k: safe(v) if k not in ("periodo",) else v for k,v in dict(x).items()} for x in pp]
    })

# ═══════════════════════════════════════════════════════════════
# PDF REAL
# ═══════════════════════════════════════════════════════════════
@app.route("/api/exportar_pdf", methods=["GET"])
@role_required("lider","admin")
def exportar_pdf():
    ini = request.args.get("data_ini","")
    fim = request.args.get("data_fim","")
    per = request.args.get("periodo","")
    tpc = request.args.get("tipo_culto","")
    culto_id = request.args.get("culto_id","")

    sql = "SELECT id,data,hora,dia_semana,periodo,COALESCE(tipo_culto,'Culto Regular') as tipo_culto,responsavel,presentes,visitantes,criancas,observacoes FROM cultos WHERE 1=1"; p=[]
    if culto_id: sql+=" AND id=?"; p.append(culto_id)
    if ini: sql+=" AND data>=?"; p.append(ini)
    if fim: sql+=" AND data<=?"; p.append(fim)
    if per: sql+=" AND periodo=?"; p.append(per)
    if tpc: sql+=" AND tipo_culto=?"; p.append(tpc)
    sql+=" ORDER BY data ASC"

    with get_db() as conn:
        cultos = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]

    total_p = sum(c["presentes"]  for c in cultos)
    total_v = sum(c["visitantes"] for c in cultos)
    total_c = sum(c["criancas"]   for c in cultos)
    n = max(len(cultos),1)
    media_p = round(total_p/n,1)
    media_v = round(total_v/n,1)

    # Cards de culto — um por linha, estilo mobile
    cards_html = ""
    for c in cultos:
        periodo_cor = {"Manhã":"#F59E0B","Tarde":"#3B82F6","Noite":"#6366F1"}.get(c["periodo"],"#64748B")
        cards_html += f"""
        <div class="culto-card">
          <div class="culto-header">
            <div class="culto-data">
              <div class="culto-dia-num">{c["data"].split("-")[2] if c.get("data") else ""}</div>
              <div class="culto-mes">{["","Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"][int(c["data"].split("-")[1])] if c.get("data") else ""}</div>
            </div>
            <div class="culto-info">
              <div class="culto-semana">{c["dia_semana"]}</div>
              <div class="culto-tipo">{c["tipo_culto"] or "Culto Regular"}</div>
              <div class="culto-resp">{c["responsavel"]}</div>
            </div>
            <span class="periodo-badge" style="background:{periodo_cor}">{c["periodo"]}</span>
          </div>
          <div class="culto-nums">
            <div class="cn"><div class="cn-v">{c["presentes"]}</div><div class="cn-l">Presentes</div></div>
            <div class="cn"><div class="cn-v" style="color:#059669">{c["visitantes"]}</div><div class="cn-l">Visitantes</div></div>
            <div class="cn"><div class="cn-v" style="color:#7C3AED">{c["criancas"]}</div><div class="cn-l">Criancas</div></div>
          </div>
        </div>"""

    filtro_txt = ""
    if ini or fim: filtro_txt = f" | Periodo: {br(ini) if ini else '...'} ate {br(fim) if fim else '...'}"
    if per: filtro_txt += f" | {per}"

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Relatorio Igreja ABA</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;background:#F1F5F9;color:#1E293B;padding:0;}}
.btn-print{{display:block;width:calc(100% - 32px);margin:16px auto 0;padding:14px;background:#0A2463;color:#fff;border:none;border-radius:12px;font-size:15px;font-weight:700;letter-spacing:.5px;cursor:pointer;}}
.hint{{text-align:center;font-size:11px;color:#94A3B8;margin:6px 16px 14px;}}
.header{{background:linear-gradient(135deg,#0A2463,#1B4FA8);color:#fff;padding:20px 16px 16px;margin:0 0 16px;}}
.header-top{{display:flex;align-items:center;gap:12px;margin-bottom:10px;}}
.logo-circle{{width:44px;height:44px;background:rgba(255,255,255,.15);border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:20px;font-weight:900;flex-shrink:0;}}
.header h1{{font-size:22px;letter-spacing:3px;font-weight:900;}}
.header .sub{{font-size:11px;opacity:.7;letter-spacing:1px;}}
.header-meta{{font-size:11px;opacity:.6;margin-top:4px;}}
.stats{{display:grid;grid-template-columns:1fr 1fr;gap:10px;padding:0 16px 16px;}}
.stat{{background:#fff;border-radius:14px;padding:14px 12px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.08);}}
.stat.blue{{background:linear-gradient(135deg,#0A2463,#1B4FA8);}}
.stat.green{{background:linear-gradient(135deg,#065F46,#059669);}}
.stat .sv{{font-size:28px;font-weight:800;color:#0A2463;line-height:1;}}
.stat.blue .sv,.stat.green .sv{{color:#fff;}}
.stat .sl{{font-size:10px;color:#64748B;text-transform:uppercase;letter-spacing:.8px;margin-top:4px;font-weight:600;}}
.stat.blue .sl,.stat.green .sl{{color:rgba(255,255,255,.75);}}
.section-title{{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;color:#64748B;padding:0 16px 8px;}}
.culto-card{{background:#fff;border-radius:14px;margin:0 16px 10px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08);}}
.culto-header{{display:flex;align-items:flex-start;gap:12px;margin-bottom:12px;}}
.culto-data{{background:#F1F5F9;border-radius:10px;padding:8px 10px;text-align:center;min-width:46px;}}
.culto-dia-num{{font-size:22px;font-weight:800;color:#0A2463;line-height:1;}}
.culto-mes{{font-size:10px;color:#64748B;text-transform:uppercase;font-weight:600;}}
.culto-info{{flex:1;}}
.culto-semana{{font-size:12px;color:#64748B;}}
.culto-tipo{{font-size:14px;font-weight:700;color:#0A2463;}}
.culto-resp{{font-size:12px;color:#64748B;margin-top:2px;}}
.periodo-badge{{padding:3px 10px;border-radius:20px;font-size:10px;font-weight:700;color:#fff;white-space:nowrap;height:fit-content;margin-top:2px;}}
.culto-nums{{display:grid;grid-template-columns:repeat(3,1fr);gap:6px;background:#F8FAFF;border-radius:10px;padding:10px;}}
.cn{{text-align:center;}}
.cn-v{{font-size:20px;font-weight:800;color:#0A2463;line-height:1;}}
.cn-l{{font-size:9px;color:#94A3B8;text-transform:uppercase;margin-top:2px;font-weight:600;}}
.totais-card{{background:linear-gradient(135deg,#0A2463,#1B4FA8);border-radius:14px;margin:0 16px 16px;padding:14px;color:#fff;}}
.totais-title{{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;opacity:.7;margin-bottom:10px;}}
.totais-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;}}
.tg{{text-align:center;}}
.tg-v{{font-size:22px;font-weight:800;}}
.tg-l{{font-size:9px;opacity:.7;text-transform:uppercase;letter-spacing:.5px;}}
.footer{{text-align:center;font-size:10px;color:#94A3B8;padding:16px;}}
@media print{{
  .btn-print,.hint{{display:none;}}
  body{{background:#fff;}}
  .culto-card,.stat,.totais-card{{box-shadow:none;border:1px solid #E2E8F0;}}
}}
</style>
</head>
<body>
<button class="btn-print" onclick="window.print()">Salvar como PDF</button>
<p class="hint">No celular: toque nos 3 pontos e escolha "Imprimir"</p>

<div class="header">
  <div class="header-top">
    <div class="logo-circle">A</div>
    <div>
      <h1>IGREJA ABA</h1>
      <div class="sub">UM LAR PARA PERTENCER</div>
    </div>
  </div>
  <div class="header-meta">Relatorio de Cultos &nbsp;|&nbsp; Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}{filtro_txt}</div>
</div>

<div class="stats">
  <div class="stat blue">
    <div class="sv">{len(cultos)}</div>
    <div class="sl">Cultos</div>
  </div>
  <div class="stat green">
    <div class="sv">{total_p}</div>
    <div class="sl">Presentes</div>
  </div>
  <div class="stat">
    <div class="sv" style="color:#059669">{total_v}</div>
    <div class="sl">Visitantes</div>
  </div>
  <div class="stat">
    <div class="sv" style="color:#7C3AED">{total_c}</div>
    <div class="sl">Criancas</div>
  </div>
</div>

<div class="section-title">Registros ({len(cultos)} cultos)</div>
{cards_html}

<div class="totais-card">
  <div class="totais-title">Totais & Medias</div>
  <div class="totais-grid">
    <div class="tg"><div class="tg-v">{total_p}</div><div class="tg-l">Total Pres.</div></div>
    <div class="tg"><div class="tg-v">{total_v}</div><div class="tg-l">Total Visit.</div></div>
    <div class="tg"><div class="tg-v">{total_c}</div><div class="tg-l">Total Crian.</div></div>
    <div class="tg"><div class="tg-v">{media_p}</div><div class="tg-l">Media Pres.</div></div>
    <div class="tg"><div class="tg-v">{media_v}</div><div class="tg-l">Media Visit.</div></div>
    <div class="tg"><div class="tg-v">{n}</div><div class="tg-l">Cultos</div></div>
  </div>
</div>

<div class="footer">Igreja ABA - Um Lar Para Pertencer</div>
</body></html>"""

    resp = make_response(html)
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    return resp
@app.route("/api/exportar_excel", methods=["GET"])
@login_required
def exportar_excel():
    ini = request.args.get("data_ini","")
    fim = request.args.get("data_fim","")
    per = request.args.get("periodo","")
    tpc = request.args.get("tipo_culto","")
    culto_id = request.args.get("culto_id","")
    with get_db() as conn:
        sql = "SELECT * FROM cultos WHERE 1=1"
        p = []
        if culto_id: sql+=" AND id=?"; p.append(culto_id)
        if ini: sql+=" AND data>=?"; p.append(ini)
        if fim: sql+=" AND data<=?"; p.append(fim)
        if per: sql+=" AND periodo=?"; p.append(per)
        if tpc: sql+=" AND tipo_culto=?"; p.append(tpc)
        sql += " ORDER BY data ASC"
        cultos = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]
        # Relatórios de GC
        gc_rows = [dict(r) for r in conn.execute(
            "SELECT * FROM relatorios_gc ORDER BY dia ASC"
        ).fetchall()]

    wb = openpyxl.Workbook()

    # ── Estilos ────────────────────────────────────────
    azul     = PatternFill("solid", fgColor="0A2463")
    azul2    = PatternFill("solid", fgColor="1B4FA8")
    cinza    = PatternFill("solid", fgColor="EBF5FF")
    verde    = PatternFill("solid", fgColor="D1FAE5")
    amarelo  = PatternFill("solid", fgColor="FEF3C7")
    branco   = PatternFill("solid", fgColor="FFFFFF")
    txt_branco = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    txt_azul   = Font(color="0A2463", bold=True, name="Calibri", size=11)
    txt_normal = Font(name="Calibri", size=10)
    txt_bold   = Font(name="Calibri", size=10, bold=True)
    centro  = Alignment(horizontal="center", vertical="center")
    esq     = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borda   = Border(
        left=Side(style="thin",color="D1D5DB"),
        right=Side(style="thin",color="D1D5DB"),
        top=Side(style="thin",color="D1D5DB"),
        bottom=Side(style="thin",color="D1D5DB")
    )

    def estilizar_cabecalho(ws, row, colunas, fill=None):
        fill = fill or azul
        for col_idx, titulo in enumerate(colunas, 1):
            c = ws.cell(row=row, column=col_idx, value=titulo)
            c.fill = fill; c.font = txt_branco; c.alignment = centro; c.border = borda

    def estilizar_celula(ws, row, col, valor, bold=False, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=valor)
        c.fill = fill or branco
        c.font = txt_bold if bold else txt_normal
        c.alignment = align or centro
        c.border = borda
        return c

    # ════════════════════════════════════════════════════
    # ABA 1: Resumo Geral
    # ════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "Resumo Geral"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 18
    ws1.row_dimensions[1].height = 35

    # Título
    ws1.merge_cells("A1:B1")
    c = ws1["A1"]; c.value = "IGREJA ABA — Relatório de Cultos"
    c.fill = azul; c.font = Font(color="FFFFFF", bold=True, name="Calibri", size=14)
    c.alignment = centro

    dados_resumo = [
        ("Total de Cultos",           len(cultos)),
        ("Total de Presentes",         sum(c["presentes"] for c in cultos)),
        ("Total de Visitantes",        sum(c["visitantes"] for c in cultos)),
        ("Total de Crianças",          sum(c["criancas"] for c in cultos)),
        ("Média de Presentes/Culto",   round(sum(c["presentes"] for c in cultos)/max(len(cultos),1),1)),
        ("Média de Visitantes/Culto",  round(sum(c["visitantes"] for c in cultos)/max(len(cultos),1),1)),
        ("Média de Crianças/Culto",    round(sum(c["criancas"] for c in cultos)/max(len(cultos),1),1)),
    ]
    for i, (label, valor) in enumerate(dados_resumo, 3):
        ws1.row_dimensions[i].height = 22
        estilizar_celula(ws1, i, 1, label, bold=True, fill=cinza if i%2==0 else branco, align=esq)
        estilizar_celula(ws1, i, 2, valor, bold=True, fill=cinza if i%2==0 else branco)

    # Por período
    ws1["A11"] = "POR PERÍODO"
    ws1["A11"].fill = azul2; ws1["A11"].font = txt_branco; ws1["A11"].alignment = esq
    ws1.merge_cells("A11:B11")
    estilizar_cabecalho(ws1, 12, ["Período","Total Presentes"])
    periodos = {}
    for c in cultos:
        p = c.get("periodo","—")
        if p not in periodos: periodos[p] = 0
        periodos[p] += c["presentes"]
    for i,(per,tot) in enumerate(periodos.items(),13):
        estilizar_celula(ws1,i,1,per,align=esq)
        estilizar_celula(ws1,i,2,tot,bold=True)

    # ════════════════════════════════════════════════════
    # ABA 2: Todos os Cultos
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Cultos")
    colunas2 = ["Data","Dia","Período","Tipo","Responsável","Presentes","Visitantes","Crianças","Observações"]
    larguras2 = [12,10,8,18,20,10,10,10,30]
    for i,(c,l) in enumerate(zip(colunas2,larguras2),1):
        ws2.column_dimensions[get_column_letter(i)].width = l
    ws2.row_dimensions[1].height = 28
    estilizar_cabecalho(ws2, 1, colunas2)

    for i,c in enumerate(cultos,2):
        fill = cinza if i%2==0 else branco
        ws2.row_dimensions[i].height = 20
        vals = [br(c["data"]),c.get("dia_semana",""),c.get("periodo",""),
                c.get("tipo_culto","Culto Regular"),c.get("responsavel",""),
                c["presentes"],c["visitantes"],c["criancas"],c.get("observacoes","")]
        for j,v in enumerate(vals,1):
            al = esq if j in (5,9) else centro
            estilizar_celula(ws2,i,j,v,fill=fill,align=al)

    # Linha de totais
    tot_row = len(cultos)+2
    ws2.row_dimensions[tot_row].height = 24
    estilizar_celula(ws2,tot_row,1,"TOTAIS",bold=True,fill=amarelo)
    ws2.merge_cells(f"A{tot_row}:E{tot_row}")
    estilizar_celula(ws2,tot_row,6,sum(c["presentes"] for c in cultos),bold=True,fill=amarelo)
    estilizar_celula(ws2,tot_row,7,sum(c["visitantes"] for c in cultos),bold=True,fill=amarelo)
    estilizar_celula(ws2,tot_row,8,sum(c["criancas"] for c in cultos),bold=True,fill=amarelo)

    media_row = tot_row+1
    ws2.row_dimensions[media_row].height = 24
    estilizar_celula(ws2,media_row,1,"MÉDIAS/CULTO",bold=True,fill=verde)
    ws2.merge_cells(f"A{media_row}:E{media_row}")
    n = max(len(cultos),1)
    estilizar_celula(ws2,media_row,6,round(sum(c["presentes"] for c in cultos)/n,1),bold=True,fill=verde)
    estilizar_celula(ws2,media_row,7,round(sum(c["visitantes"] for c in cultos)/n,1),bold=True,fill=verde)
    estilizar_celula(ws2,media_row,8,round(sum(c["criancas"] for c in cultos)/n,1),bold=True,fill=verde)

    # ════════════════════════════════════════════════════
    # ABA 3: Relatórios de GC
    # ════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Relatórios GC")
    colunas3 = ["Data","GC","Líder","Anfitrião","Membros","Visitantes","Líder Trein.","Nome Líder Trein.","Observações"]
    larguras3 = [12,22,20,20,10,10,12,22,35]
    for i,(c,l) in enumerate(zip(colunas3,larguras3),1):
        ws3.column_dimensions[get_column_letter(i)].width = l
    ws3.row_dimensions[1].height = 28
    estilizar_cabecalho(ws3, 1, colunas3)

    for i,r in enumerate(gc_rows,2):
        fill = cinza if i%2==0 else branco
        ws3.row_dimensions[i].height = 20
        vals = [br(r.get("dia","")),r.get("gc_nome",""),r.get("lider_nome",""),
                r.get("anfitriao",""),r.get("membros_presentes",0),r.get("visitantes",0),
                "Sim" if r.get("lider_treinamento") else "Não",
                r.get("nome_lider_trein",""),r.get("observacoes","")]
        for j,v in enumerate(vals,1):
            al = esq if j in (2,3,4,8,9) else centro
            estilizar_celula(ws3,i,j,v,fill=fill,align=al)

    # ════════════════════════════════════════════════════
    # ABA 4: Ranking GCs
    # ════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Ranking GCs")
    ws4.column_dimensions["A"].width = 25
    for col in ["B","C","D","E","F"]: ws4.column_dimensions[col].width = 14
    ws4.row_dimensions[1].height = 28
    estilizar_cabecalho(ws4, 1, ["GC","Reuniões","Total Membros","Total Visit.","Média Membros","Média Visit."])

    gc_stats = {}
    for r in gc_rows:
        nome = r.get("gc_nome","—")
        if nome not in gc_stats: gc_stats[nome] = {"reunioes":0,"membros":0,"visitantes":0}
        gc_stats[nome]["reunioes"]   += 1
        gc_stats[nome]["membros"]    += r.get("membros_presentes",0)
        gc_stats[nome]["visitantes"] += r.get("visitantes",0)

    ranking = sorted(gc_stats.items(), key=lambda x: x[1]["membros"], reverse=True)
    for i,(nome,s) in enumerate(ranking,2):
        fill = cinza if i%2==0 else branco
        ws4.row_dimensions[i].height = 22
        n_r = max(s["reunioes"],1)
        vals = [nome,s["reunioes"],s["membros"],s["visitantes"],
                round(s["membros"]/n_r,1),round(s["visitantes"]/n_r,1)]
        for j,v in enumerate(vals,1):
            al = esq if j==1 else centro
            estilizar_celula(ws4,i,j,v,bold=(j==1),fill=fill,align=al)

    # ════════════════════════════════════════════════════
    # ABA 5: Por Período (média separada por período/tipo)
    # ════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Por Período")
    ws5.column_dimensions["A"].width = 22
    for col in ["B","C","D","E","F"]: ws5.column_dimensions[col].width = 14
    ws5.row_dimensions[1].height = 28
    estilizar_cabecalho(ws5, 1, ["Período","Qtd. Cultos","Total Presentes","Total Visitantes","Média Presentes","Média Visitantes"])

    # Agrupa por período
    por_periodo = {}
    for c in cultos:
        p = c.get("periodo","—") or "—"
        if p not in por_periodo: por_periodo[p] = {"qtd":0,"presentes":0,"visitantes":0,"criancas":0}
        por_periodo[p]["qtd"] += 1
        por_periodo[p]["presentes"] += c["presentes"]
        por_periodo[p]["visitantes"] += c["visitantes"]
        por_periodo[p]["criancas"] += c["criancas"]

    for i,(per,s) in enumerate(sorted(por_periodo.items()),2):
        fill = cinza if i%2==0 else branco
        n_p = max(s["qtd"],1)
        estilizar_celula(ws5,i,1,per,bold=True,fill=fill,align=esq)
        estilizar_celula(ws5,i,2,s["qtd"],fill=fill)
        estilizar_celula(ws5,i,3,s["presentes"],fill=fill)
        estilizar_celula(ws5,i,4,s["visitantes"],fill=fill)
        estilizar_celula(ws5,i,5,round(s["presentes"]/n_p,1),bold=True,fill=fill)
        estilizar_celula(ws5,i,6,round(s["visitantes"]/n_p,1),bold=True,fill=fill)

    # Separador
    sep = len(por_periodo)+3
    ws5.merge_cells(f"A{sep}:F{sep}")
    c = ws5.cell(row=sep,column=1,value="POR TIPO DE CULTO")
    c.fill=azul2; c.font=Font(color="FFFFFF",bold=True,name="Calibri",size=11); c.alignment=centro

    estilizar_cabecalho(ws5, sep+1, ["Tipo de Culto","Qtd. Cultos","Total Presentes","Total Visitantes","Média Presentes","Média Visitantes"], fill=azul2)

    por_tipo = {}
    for c in cultos:
        t = c.get("tipo_culto","Culto Regular") or "Culto Regular"
        if t not in por_tipo: por_tipo[t] = {"qtd":0,"presentes":0,"visitantes":0}
        por_tipo[t]["qtd"] += 1
        por_tipo[t]["presentes"] += c["presentes"]
        por_tipo[t]["visitantes"] += c["visitantes"]

    for i,(tp,s) in enumerate(sorted(por_tipo.items(),key=lambda x:x[1]["presentes"],reverse=True), sep+2):
        fill = cinza if i%2==0 else branco
        n_t = max(s["qtd"],1)
        estilizar_celula(ws5,i,1,tp,bold=True,fill=fill,align=esq)
        estilizar_celula(ws5,i,2,s["qtd"],fill=fill)
        estilizar_celula(ws5,i,3,s["presentes"],fill=fill)
        estilizar_celula(ws5,i,4,s["visitantes"],fill=fill)
        estilizar_celula(ws5,i,5,round(s["presentes"]/n_t,1),bold=True,fill=fill)
        estilizar_celula(ws5,i,6,round(s["visitantes"]/n_t,1),bold=True,fill=fill)

    # ════════════════════════════════════════════════════
    # ABA 6: Visão Consolidada
    # ════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Visão Consolidada")
    ws6.column_dimensions["A"].width = 30; ws6.column_dimensions["B"].width = 18
    ws6.row_dimensions[1].height = 35
    ws6.merge_cells("A1:B1")
    c=ws6["A1"]; c.value="VISÃO CONSOLIDADA — IGREJA ABA"
    c.fill=azul; c.font=Font(color="FFFFFF",bold=True,name="Calibri",size=14); c.alignment=centro

    n_all = max(len(cultos),1)
    total_p_all = sum(c["presentes"] for c in cultos)
    total_v_all = sum(c["visitantes"] for c in cultos)
    total_c_all = sum(c["criancas"] for c in cultos)
    total_gc_relat = len(gc_rows)
    total_gc_membros = sum(r.get("membros_presentes",0) for r in gc_rows)

    dados_consol = [
        ("═══ CULTOS ═══", ""),
        ("Total de Cultos", len(cultos)),
        ("Total de Presentes", total_p_all),
        ("Total de Visitantes", total_v_all),
        ("Total de Crianças", total_c_all),
        ("Média de Presentes/Culto", round(total_p_all/n_all,1)),
        ("Média de Visitantes/Culto", round(total_v_all/n_all,1)),
        ("Média de Crianças/Culto", round(total_c_all/n_all,1)),
        ("═══ GCS ═══", ""),
        ("Total de Relatórios GC", total_gc_relat),
        ("Total de Membros GC", total_gc_membros),
        ("Média Membros/Reunião", round(total_gc_membros/max(total_gc_relat,1),1)),
        ("GCs com Relatório", len(set(r.get("gc_nome","") for r in gc_rows))),
    ]
    for i,(label,valor) in enumerate(dados_consol,3):
        ws6.row_dimensions[i].height = 22
        is_section = str(valor)==""
        fill_cell = azul2 if is_section else (cinza if i%2==0 else branco)
        font_cell = Font(color="FFFFFF",bold=True,name="Calibri",size=11) if is_section else Font(name="Calibri",size=10,bold=True)
        c1=ws6.cell(row=i,column=1,value=label); c1.fill=fill_cell; c1.font=font_cell; c1.alignment=esq; c1.border=borda
        c2=ws6.cell(row=i,column=2,value=valor if not is_section else ""); c2.fill=fill_cell; c2.font=font_cell; c2.alignment=centro; c2.border=borda
        if is_section: ws6.merge_cells(f"A{i}:B{i}")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"igrejaaba_{date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@app.route("/api/gcs/direcionamentos/<int:did>", methods=["DELETE"])
@role_required("admin","lider")
def deletar_direcionamento(did):
    with get_db() as conn:
        conn.execute(qmark("DELETE FROM gc_direcionamentos WHERE id=?"), (did,))
        conn.commit()
    return jsonify({"ok":True})


@app.route("/api/gcs/resetar", methods=["POST"])
@role_required("admin")
def resetar_gcs():
    """Apaga todos os GCs e reinsere a lista padrão — resolve duplicatas e dados errados"""
    with get_db() as conn:
        try:
            conn.execute("DELETE FROM gc_direcionamentos")
            conn.execute("DELETE FROM grupos_crescimento")
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"Erro ao limpar GCs: {e}")
            return jsonify({"erro": str(e)}), 500
    if USE_PG:
        _init_pg_gcs()
    else:
        _init_sqlite_gcs()
    return jsonify({"ok": True, "msg": f"{len(_GCS)} GCs reinseridos"})

def _init_pg_gcs():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    cur = conn.cursor()
    for gc in _GCS:
        cur.execute(
            qmark("INSERT INTO grupos_crescimento(nome,lider,endereco,bairro,cidade,setor,cor_hex,lat,lng) VALUES(?,?,?,?,?,?,?,?,?)"),
            gc
        )
    conn.commit(); conn.close()

def _init_sqlite_gcs():
    conn = sqlite3.connect(DB_PATH)
    for gc in _GCS:
        conn.execute(
            "INSERT OR IGNORE INTO grupos_crescimento(nome,lider,endereco,bairro,cidade,setor,cor_hex,lat,lng) VALUES(?,?,?,?,?,?,?,?,?)",
            gc
        )
    conn.commit(); conn.close()

@app.route("/api/cameras", methods=["GET"])
@login_required
def listar_cameras():
    with get_db() as conn:
        rows=conn.execute("SELECT * FROM cameras WHERE ativa=1 ORDER BY nome").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/cameras", methods=["POST"])
@role_required("admin")
def criar_camera():
    d=request.get_json(force=True) or {}
    nome=d.get("nome","").strip(); url=d.get("url","").strip()
    if not nome: return jsonify({"erro":"Nome obrigatório"}),400
    with get_db() as conn:
        cur=conn.execute(qmark("INSERT INTO cameras(nome,url,local) VALUES(?,?,?)"), (nome,url,d.get("local","")))
        conn.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/cameras/<int:cid>", methods=["DELETE"])
@role_required("admin")
def del_camera(cid):
    with get_db() as conn:
        conn.execute(qmark("UPDATE cameras SET ativa=0 WHERE id=?"), (cid,)); conn.commit()
    return jsonify({"ok":True})

@app.route("/api/contagem/sessoes", methods=["GET"])
@login_required
def listar_sessoes():
    with get_db() as conn:
        rows=conn.execute("""SELECT s.*,c.data as culto_data,c.periodo FROM contagem_sessoes s
            LEFT JOIN cultos c ON c.id=s.culto_id ORDER BY s.iniciado_em DESC LIMIT 50""").fetchall()
    return jsonify([{**dict(r),"culto_data_br":br(r["culto_data"]) if r["culto_data"] else ""} for r in rows])

@app.route("/api/contagem/sessoes", methods=["POST"])
@login_required
def criar_sessao():
    d=request.get_json(force=True) or {}
    with get_db() as conn:
        cur=conn.execute(qmark("INSERT INTO contagem_sessoes(culto_id,camera_id,camera_nome) VALUES(?,?,?)"), (d.get("culto_id"),d.get("camera_id"),d.get("camera_nome","")))
        conn.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/contagem/sessoes/<int:sid>/encerrar", methods=["POST"])
@login_required
def encerrar_sessao(sid):
    with get_db() as conn:
        conn.execute(qmark("UPDATE contagem_sessoes SET status='encerrada',encerrado_em=? WHERE id=?"), (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), sid,)); conn.commit()
    return jsonify({"ok":True})

@app.route("/api/contagem/registrar", methods=["POST"])
@login_required
def registrar_passagem():
    d=request.get_json(force=True) or {}
    sid=d.get("sessao_id"); tid=d.get("track_id"); direcao=d.get("direcao")
    if not sid or tid is None or not direcao:
        return jsonify({"erro":"sessao_id, track_id e direcao obrigatórios"}),400
    with get_db() as conn:
        existe=conn.execute(qmark("SELECT id FROM contagem_registros WHERE sessao_id=? AND track_id=? AND direcao=?"), (sid,tid,direcao)).fetchone()
        if existe: return jsonify({"ok":True,"duplicado":True})
        conn.execute(qmark("INSERT INTO contagem_registros(sessao_id,track_id,direcao,confianca) VALUES(?,?,?,?)"), (sid,tid,direcao,d.get("confianca",1.0)))
        col="total_entradas" if direcao=="entrada" else "total_saidas"
        conn.execute(f"UPDATE contagem_sessoes SET {col}={col}+1 WHERE id=?",(sid,))
        conn.commit()
    return jsonify({"ok":True,"duplicado":False})

@app.route("/api/contagem/tempo_real/<int:sid>", methods=["GET"])
@login_required
def tempo_real(sid):
    with get_db() as conn:
        s=conn.execute(qmark("SELECT * FROM contagem_sessoes WHERE id=?"), (sid,)).fetchone()
        if not s: return jsonify({"erro":"Sessão não encontrada"}),404
    s=dict(s)
    return jsonify({"sessao_id":sid,"entradas":s["total_entradas"],"saidas":s["total_saidas"],
                    "dentro_agora":max(0,s["total_entradas"]-s["total_saidas"]),"status":s["status"]})






# ═══════════════════════════════════════════════════════════════
# VOLUNTÁRIOS
# ═══════════════════════════════════════════════════════════════

@app.route("/api/voluntarios", methods=["GET"])
@login_required
def listar_voluntarios():
    depto_filtro = get_depto_usuario()
    todos = request.args.get("todos","")
    cargo = session.get("usuario_cargo","")
    if cargo in ("admin","lider") or todos:
        depto_filtro = None
    with get_db() as conn:
        if depto_filtro:
            rows = [dict(r) for r in conn.execute(
                qmark("""SELECT v.*, COALESCE(d.icone||' '||d.nome,'') as departamento_nome
                         FROM voluntarios v LEFT JOIN departamentos d ON d.id=v.departamento_id
                         WHERE v.ativo=1 AND v.departamento_id=? ORDER BY v.nome"""),
                (depto_filtro,)
            ).fetchall()]
        else:
            rows = [dict(r) for r in conn.execute(
                """SELECT v.*, COALESCE(d.icone||' '||d.nome,'') as departamento_nome
                   FROM voluntarios v LEFT JOIN departamentos d ON d.id=v.departamento_id
                   WHERE v.ativo=1 ORDER BY d.ordem, v.nome"""
            ).fetchall()]
    return jsonify(rows)


@app.route("/api/voluntarios/por_departamento/<int:did>", methods=["GET"])
@login_required
def voluntarios_por_depto(did):
    with get_db() as conn:
        # Busca nome do departamento para também casar pelo campo texto 'departamentos'
        dep = conn.execute(qmark("SELECT nome FROM departamentos WHERE id=?"), (did,)).fetchone()
        dep_nome = dep["nome"] if dep else ""
        rows = [dict(r) for r in conn.execute(
            qmark("""SELECT id,nome,telefone FROM voluntarios
                     WHERE ativo=1 AND (departamento_id=? OR (departamento_id IS NULL AND departamentos LIKE ?))
                     ORDER BY nome"""),
            (did, f"%{dep_nome}%")
        ).fetchall()]
    return jsonify(rows)

@app.route("/api/voluntarios", methods=["POST"])
@role_required("admin","lider","lider_depto")
def criar_voluntario():
    d = request.get_json(force=True) or {}
    nome = d.get("nome","").strip()
    tel  = d.get("telefone","").strip()
    if not nome or not tel:
        return jsonify({"erro":"Nome e telefone são obrigatórios"}),400
    depto_id = d.get("departamento_id") or get_depto_usuario()
    try:
        with get_db() as conn:
            cur = conn.execute(
                qmark("INSERT INTO voluntarios(nome,telefone,departamentos,departamento_id) VALUES(?,?,?,?)"),
                (nome, tel, d.get("departamentos",""), depto_id)
            )
            conn.commit()
        return jsonify({"ok":True,"id":cur.lastrowid})
    except Exception as e:
        logger.error(f"Erro ao criar voluntário: {e}")
        return jsonify({"erro":f"Erro ao salvar: {str(e)}"}),500

@app.route("/api/voluntarios/<int:vid>", methods=["PUT"])
@role_required("admin","lider")
def editar_voluntario(vid):
    d = request.get_json(force=True) or {}
    with get_db() as conn:
        v = conn.execute(qmark("SELECT * FROM voluntarios WHERE id=?"), (vid,)).fetchone()
        if not v: return jsonify({"erro":"Não encontrado"}),404
        conn.execute(
            qmark("UPDATE voluntarios SET nome=?,telefone=?,departamentos=? WHERE id=?"),
            (d.get("nome",v["nome"]), d.get("telefone",v["telefone"]),
             d.get("departamentos",v["departamentos"]), vid)
        )
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/voluntarios/<int:vid>", methods=["DELETE"])
@role_required("admin","lider")
def del_voluntario(vid):
    with get_db() as conn:
        conn.execute(qmark("UPDATE voluntarios SET ativo=0 WHERE id=?"), (vid,))
        conn.commit()
    return jsonify({"ok":True})

# ═══════════════════════════════════════════════════════════════
# ESCALA — PUBLICAÇÃO E CONFIRMAÇÕES
# ═══════════════════════════════════════════════════════════════

@app.route("/api/escala/publicar", methods=["POST"])
@role_required("admin","lider","lider_depto")
def publicar_escala():
    d = request.get_json(force=True) or {}
    mes = d.get("mes","").strip()
    if not mes: return jsonify({"erro":"Mês obrigatório"}),400
    with get_db() as conn:
        # Registra publicação
        existe = conn.execute(qmark("SELECT id FROM escala_publicacoes WHERE mes=?"), (mes,)).fetchone()
        if existe:
            conn.execute(
                qmark("UPDATE escala_publicacoes SET status='publicada',publicado_em=?,publicado_por=? WHERE mes=?"),
                (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get("usuario_nome","?"), mes)
            )
        else:
            conn.execute(
                qmark("INSERT INTO escala_publicacoes(mes,status,publicado_em,publicado_por) VALUES(?,?,?,?)"),
                (mes,"publicada", datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get("usuario_nome","?"))
            )
        conn.commit()
    return jsonify({"ok":True,"msg":f"Escala de {mes} publicada!"})

@app.route("/api/escala/reabrir", methods=["POST"])
@role_required("admin","lider")
def reabrir_escala():
    d = request.get_json(force=True) or {}
    mes = d.get("mes","").strip()
    with get_db() as conn:
        conn.execute(qmark("UPDATE escala_publicacoes SET status='rascunho' WHERE mes=?"), (mes,))
        conn.commit()
    return jsonify({"ok":True,"msg":f"Escala de {mes} reaberta para edição"})

@app.route("/api/escala/status/<mes>", methods=["GET"])
@login_required
def status_escala(mes):
    with get_db() as conn:
        pub = conn.execute(qmark("SELECT * FROM escala_publicacoes WHERE mes=?"), (mes,)).fetchone()
    if not pub:
        return jsonify({"status":"rascunho","publicado_em":None,"publicado_por":""})
    return jsonify(dict(pub))

@app.route("/api/escala/notificar", methods=["POST"])
@role_required("admin","lider","lider_depto")
def notificar_voluntarios():
    """Gera links WhatsApp para notificar TODOS os voluntários de uma vez"""
    d = request.get_json(force=True) or {}
    mes = d.get("mes","").strip()
    publicar_primeiro = d.get("publicar", False)
    if not mes: return jsonify({"erro":"Mês obrigatório"}),400
    # Publica automaticamente ao notificar se solicitado
    if publicar_primeiro:
        try:
            with get_db() as conn:
                existe = conn.execute(qmark("SELECT id FROM escala_publicacoes WHERE mes=?"), (mes,)).fetchone()
                if existe:
                    conn.execute(qmark("UPDATE escala_publicacoes SET status='publicada',publicado_em=?,publicado_por=? WHERE mes=?"),
                        (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get("usuario_nome","?"), mes))
                else:
                    conn.execute(qmark("INSERT INTO escala_publicacoes(mes,status,publicado_em,publicado_por) VALUES(?,?,?,?)"),
                        (mes,"publicada",datetime.now().strftime('%Y-%m-%d %H:%M:%S'),session.get("usuario_nome","?")))
                conn.commit()
        except Exception as e:
            logger.warning(f"Erro ao publicar: {e}")

    with get_db() as conn:
        # Busca todos os itens do mês com responsável preenchido
        itens = [dict(r) for r in conn.execute(qmark(
            """SELECT e.responsavel, e.culto_data, e.culto_periodo, d.nome as depto
               FROM escala_itens e
               JOIN departamentos d ON d.id=e.departamento_id
               WHERE e.culto_data LIKE ? AND e.responsavel!=''
               ORDER BY e.culto_data, d.ordem"""
        ), (f"{mes}%",)).fetchall()]
        # Busca voluntários para cruzar com telefone
        voluntarios = {v["nome"].lower().strip(): dict(v) for v in
                      conn.execute("SELECT * FROM voluntarios WHERE ativo=1").fetchall()}

    if not itens:
        return jsonify({"erro":"Nenhum item preenchido na escala deste mês"}),400

    # Agrupa por responsável
    por_pessoa = {}
    for item in itens:
        nome = item["responsavel"].strip()
        if nome not in por_pessoa:
            por_pessoa[nome] = []
        por_pessoa[nome].append(item)

    base = get_base()
    notificacoes = []

    for nome, escalas in por_pessoa.items():
        # Monta mensagem
        linhas = [f"Ola {nome}!", "", "Voce esta na escala da Igreja ABA:", ""]
        for e in escalas:
            data_br = br(e["culto_data"])
            linhas.append(f"- {data_br} ({e['culto_periodo']}): {e['depto']}")

        # Busca voluntário para token de confirmação
        vol = voluntarios.get(nome.lower())
        token = None
        if vol:
            token = secrets.token_urlsafe(16)
            try:
                with get_db() as conn:
                    # Gera confirmação para cada item
                    for e in escalas:
                        tk = secrets.token_urlsafe(16)
                        conn.execute(
                            qmark("""INSERT INTO escala_confirmacoes
                                (voluntario_id,voluntario_nome,culto_data,culto_periodo,departamento,token)
                                VALUES(?,?,?,?,?,?)
                                ON CONFLICT(token) DO NOTHING"""),
                            (vol["id"],nome,e["culto_data"],e["culto_periodo"],e["depto"],tk)
                        )
                    conn.commit()
                # Token geral para ver tudo de uma vez
                token_geral = secrets.token_urlsafe(20)
                linhas.append("")
                linhas.append(f"Para confirmar ou solicitar troca:")
                linhas.append(f"{base}/confirmar/{token_geral}?nome={urllib.parse.quote(nome)}&mes={mes}")
            except Exception as e:
                logger.warning(f"Erro gerando token: {e}")
        else:
            linhas.append("")
            linhas.append("Qualquer duvida fale com o coordenador.")

        linhas += ["","Deus abencoe!","Igreja ABA"]
        msg = "\n".join(linhas)

        # Monta link WhatsApp
        tel = ""
        if vol:
            tel = re.sub(r"\D","",vol["telefone"])
            if not tel.startswith("55"): tel = "55" + tel
        wa_link = f"https://wa.me/{tel}?text={urllib.parse.quote(msg)}" if tel else None

        notificacoes.append({
            "nome": nome,
            "telefone": vol["telefone"] if vol else None,
            "tem_telefone": bool(vol),
            "qtd_escalas": len(escalas),
            "wa_link": wa_link,
            "msg_preview": "\n".join(linhas[:6]) + "..."
        })

    return jsonify({"ok":True,"notificacoes":notificacoes,"total":len(notificacoes)})

@app.route("/confirmar/<token>")
def pagina_confirmacao(token):
    nome = request.args.get("nome","")
    mes  = request.args.get("mes","")
    return render_template("escala_confirmacao.html", token=token, nome=nome, mes=mes)

@app.route("/api/escala/confirmacoes/<token>", methods=["GET"])
def ver_confirmacoes(token):
    """Rota pública — voluntário vê seus compromissos"""
    nome = request.args.get("nome","")
    mes  = request.args.get("mes","")
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(
            qmark("SELECT * FROM escala_confirmacoes WHERE voluntario_nome=? ORDER BY culto_data"),
            (nome,)
        ).fetchall()]
    for r in rows:
        r["data_br"] = br(r["culto_data"])
    return jsonify({"nome":nome,"mes":mes,"escalas":rows})

@app.route("/api/escala/responder/<int:cid>", methods=["POST"])
def responder_confirmacao(cid):
    """Voluntário aceita ou recusa"""
    d = request.get_json(force=True) or {}
    status = d.get("status","")
    sugestao = d.get("sugestao_troca","").strip()
    if status not in ("confirmado","recusado"):
        return jsonify({"erro":"Status inválido"}),400
    with get_db() as conn:
        conn.execute(
            qmark("UPDATE escala_confirmacoes SET status=?,sugestao_troca=?,respondido_em=? WHERE id=?"),
            (status, sugestao, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), cid)
        )
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/escala/confirmacoes_admin", methods=["GET"])
@role_required("admin","lider","lider_depto")
def ver_confirmacoes_admin():
    mes = request.args.get("mes","")
    with get_db() as conn:
        sql = """SELECT c.*, v.telefone FROM escala_confirmacoes c
                 LEFT JOIN voluntarios v ON v.id=c.voluntario_id
                 WHERE 1=1"""
        p = []
        if mes:
            sql += " AND c.culto_data LIKE ?".replace("?","%s" if USE_PG else "?")
            p.append(f"{mes}%")
        sql += " ORDER BY c.culto_data, c.status"
        rows = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]
    for r in rows:
        r["data_br"] = br(r["culto_data"])
    return jsonify(rows)


@app.route("/api/escala/minha", methods=["GET"])
@login_required
def minha_escala():
    """Retorna a escala do voluntário logado — todos os deptos onde está"""
    mes = request.args.get("mes", datetime.now().strftime("%Y-%m"))
    uid = session["usuario_id"]
    try:
        with get_db() as conn:
            # Busca o voluntário pelo usuario_id
            vol = conn.execute(
                qmark("SELECT id,nome FROM voluntarios WHERE usuario_id=? AND ativo=1"), (uid,)
            ).fetchone()
            if not vol:
                return jsonify({"erro":"Voluntário não encontrado","nome":session.get("usuario_nome",""),"mes":mes,"itens":[]})
            vol_nome = vol["nome"]
            # Busca escala pelo nome do voluntário no mês.
            # Usa LIKE no mês (texto) para evitar erros de cast em PG com datas inválidas.
            sql = """SELECT e.*, d.nome as depto_nome, d.icone as depto_icone
                     FROM escala_itens e
                     JOIN departamentos d ON d.id=e.departamento_id
                     WHERE e.responsavel=? AND e.culto_data LIKE ?
                     ORDER BY e.culto_data"""
            try:
                rows = [dict(r) for r in conn.execute(qmark(sql),(vol_nome, f"{mes}%")).fetchall()]
            except Exception as e:
                logger.warning(f"minha_escala query erro: {e}")
                rows = []
            # Busca confirmações do voluntário
            confs = {}
            try:
                for c in conn.execute(
                    qmark("SELECT culto_data,culto_periodo,departamento,status FROM escala_confirmacoes WHERE voluntario_id=?"),
                    (vol["id"],)
                ).fetchall():
                    key = f"{c['culto_data']}_{c['culto_periodo']}_{c['departamento']}"
                    confs[key] = c["status"]
            except Exception:
                pass
        for r in rows:
            r["data_br"] = br(r["culto_data"])
            key = f"{r['culto_data']}_{r['culto_periodo']}_{r['depto_nome']}"
            r["status_confirmacao"] = confs.get(key,"pendente")
        return jsonify({"nome":vol_nome,"mes":mes,"itens":rows})
    except Exception as e:
        logger.error(f"Erro minha_escala: {e}")
        return jsonify({"erro":"Erro ao carregar escala","nome":session.get("usuario_nome",""),"mes":mes,"itens":[]})

@app.route("/api/escala/verificar_conflito", methods=["POST"])
@login_required
def verificar_conflito():
    """Verifica se um voluntário já está escalado em outro depto no mesmo dia"""
    d = request.get_json(force=True) or {}
    nome       = d.get("nome","").strip()
    data       = d.get("culto_data","").strip()
    periodo    = d.get("culto_periodo","").strip()
    depto_id   = d.get("departamento_id")
    if not nome or not data: return jsonify({"conflito":False})

    with get_db() as conn:
        sql = """SELECT e.responsavel, d.nome as depto_nome
                 FROM escala_itens e
                 JOIN departamentos d ON d.id=e.departamento_id
                 WHERE e.responsavel=? AND e.culto_data=?
                 AND e.culto_periodo=? AND e.departamento_id!=?"""
        conflitos = [dict(r) for r in conn.execute(
            qmark(sql), (nome, data, periodo, depto_id or 0)
        ).fetchall()]
    if conflitos:
        deptos = ", ".join(sorted(set(c["depto_nome"] for c in conflitos)))
        return jsonify({
            "conflito": True,
            "mensagem": f"{nome} já está escalado(a) em: {deptos} neste mesmo dia e período"
        })
    return jsonify({"conflito":False})

@app.route("/api/voluntarios/<int:vid>/vincular_usuario", methods=["POST"])
@role_required("admin")
def vincular_usuario_voluntario(vid):
    """Vincula um usuário a um voluntário para acesso à escala pessoal"""
    d = request.get_json(force=True) or {}
    uid = d.get("usuario_id")
    if not uid: return jsonify({"erro":"usuario_id obrigatório"}),400
    with get_db() as conn:
        conn.execute(qmark("UPDATE voluntarios SET usuario_id=? WHERE id=?"), (uid, vid))
        conn.commit()
    return jsonify({"ok":True})

# ═══════════════════════════════════════════════════════════════
# ESCALAS
# ═══════════════════════════════════════════════════════════════


@app.route("/minha-escala")
def pagina_minha_escala():
    return render_template("minha_escala.html")

@app.route("/escala")
def pagina_escala():
    """Página pública de preenchimento de escala (para coordenadores)"""
    return render_template("escala_editor.html")

@app.route("/api/departamentos", methods=["GET"])
@login_required
def listar_departamentos():
    todos = request.args.get("todos","")
    cargo = session.get("usuario_cargo","")

    # Admin e líder geral veem todos
    if cargo in ("admin","lider"):
        with get_db() as conn:
            rows = conn.execute("SELECT * FROM departamentos WHERE ativo=1 ORDER BY ordem,nome").fetchall()
        return jsonify([dict(r) for r in rows])

    # Líder de departamento: SOMENTE o departamento dele (ignora ?todos=1)
    if cargo == "lider_depto":
        dep_id = session.get("usuario_departamento_id")
        if not dep_id:
            return jsonify([])
        with get_db() as conn:
            rows = conn.execute(
                qmark("SELECT * FROM departamentos WHERE ativo=1 AND id=? ORDER BY ordem,nome"),
                (dep_id,)
            ).fetchall()
        return jsonify([dict(r) for r in rows])

    # Voluntário de escala: SOMENTE os departamentos que escolheu no cadastro
    if cargo == "voluntario_escala":
        uid = session.get("usuario_id")
        with get_db() as conn:
            vol = conn.execute(
                qmark("SELECT departamentos FROM voluntarios WHERE usuario_id=? AND ativo=1"),
                (uid,)
            ).fetchone()
            deptos_txt = (vol["departamentos"] if vol else "") or ""
            todos_dep = conn.execute("SELECT * FROM departamentos WHERE ativo=1 ORDER BY ordem,nome").fetchall()
        nomes = [n.strip().lower() for n in deptos_txt.split(",") if n.strip()]
        filtrados = [dict(r) for r in todos_dep if r["nome"].strip().lower() in nomes]
        return jsonify(filtrados)

    # Demais: todos (fallback)
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM departamentos WHERE ativo=1 ORDER BY ordem,nome").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/departamentos", methods=["POST"])
@role_required("admin","lider")
def criar_departamento():
    d = request.get_json(force=True) or {}
    nome = d.get("nome","").strip()
    if not nome: return jsonify({"erro":"Nome obrigatório"}),400
    try:
        with get_db() as conn:
            cur = conn.execute(
                qmark("INSERT INTO departamentos(nome,icone,ordem) VALUES(?,?,?)"),
                (nome, d.get("icone","👥"), int(d.get("ordem",99)))
            )
            conn.commit()
        return jsonify({"ok":True,"id":cur.lastrowid})
    except Exception as e:
        return jsonify({"erro":"Departamento já existe" if "UNIQUE" in str(e) else str(e)}),400

@app.route("/api/departamentos/<int:did>", methods=["DELETE"])
@role_required("admin")
def del_departamento(did):
    with get_db() as conn:
        conn.execute(qmark("UPDATE departamentos SET ativo=0 WHERE id=?"), (did,))
        conn.commit()
    return jsonify({"ok":True})

# ── Itens da escala ──────────────────────────────────────────
@app.route("/api/escala", methods=["GET"])
@login_required
def listar_escala():
    mes  = request.args.get("mes","")   # YYYY-MM
    per  = request.args.get("periodo","")
    sql  = """SELECT e.*, d.nome as depto_nome, d.icone as depto_icone, d.ordem as depto_ordem
              FROM escala_itens e
              JOIN departamentos d ON d.id=e.departamento_id
              WHERE 1=1"""
    p = []
    if mes:
        if USE_PG:
            sql += " AND to_char(e.culto_data::date,'YYYY-MM')=%s"; p.append(mes)
        else:
            sql += " AND strftime('%Y-%m',e.culto_data)=?"; p.append(mes)
    if per:
        sql += " AND e.culto_periodo=?".replace("?","%s" if USE_PG else "?"); p.append(per)
    sql += " ORDER BY e.culto_data, d.ordem, d.nome"
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]
    # Mantém apenas domingos
    rows = [r for r in rows if _eh_domingo(r.get("culto_data",""))]
    for r in rows:
        r["data_br"] = br(r["culto_data"]) if r.get("culto_data") else ""
    return jsonify(rows)

@app.route("/api/escala/publica", methods=["GET"])
def escala_publica():
    """Rota pública — sem login — para coordenadores preencherem"""
    mes = request.args.get("mes","")
    per = request.args.get("periodo","")
    sql = """SELECT e.*, d.nome as depto_nome, d.icone as depto_icone, d.ordem as depto_ordem
             FROM escala_itens e
             JOIN departamentos d ON d.id=e.departamento_id
             WHERE 1=1"""
    p = []
    if mes:
        if USE_PG:
            sql += " AND to_char(e.culto_data::date,'YYYY-MM')=%s"; p.append(mes)
        else:
            sql += " AND strftime('%Y-%m',e.culto_data)=?"; p.append(mes)
    if per:
        sql += (" AND e.culto_periodo=%s" if USE_PG else " AND e.culto_periodo=?"); p.append(per)
    sql += " ORDER BY e.culto_data, d.ordem"
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]
        deptos = [dict(r) for r in conn.execute("SELECT * FROM departamentos WHERE ativo=1 ORDER BY ordem").fetchall()]
    rows = [r for r in rows if _eh_domingo(r.get("culto_data",""))]
    for r in rows: r["data_br"] = br(r["culto_data"]) if r.get("culto_data") else ""
    return jsonify({"itens": rows, "departamentos": deptos})

@app.route("/api/escala/pdf", methods=["GET"])
@role_required("admin","lider","lider_depto")
def exportar_escala_pdf():
    """Gera PDF da escala mensal organizada por departamento"""
    mes = request.args.get("mes","")
    per = request.args.get("periodo","")
    sql = """SELECT e.*, d.nome as depto_nome, d.icone as depto_icone, d.ordem as depto_ordem
             FROM escala_itens e
             JOIN departamentos d ON d.id=e.departamento_id
             WHERE e.culto_data LIKE ?"""
    p = [f"{mes}%"] if mes else ["%"]
    if per: sql += " AND e.culto_periodo=?"; p.append(per)
    sql += " ORDER BY d.ordem, d.nome, e.culto_data"
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]

    # Mantém apenas dias de culto (sexta, sábado, domingo)
    rows = [r for r in rows if _eh_domingo(r.get("culto_data",""))]

    # Monta cabeçalho do mês
    try:
        ano_m, mes_m = mes.split("-")
        MESES_BR = ["","Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
        titulo_mes = f"{MESES_BR[int(mes_m)]} de {ano_m}"
    except Exception:
        titulo_mes = mes or "Escala"

    DIAS_ABREV = {4:"SEX", 5:"SÁB", 6:"DOM"}
    def label_data(data_str):
        try:
            dd = datetime.strptime(str(data_str)[:10], "%Y-%m-%d")
            return f"{dd.strftime('%d/%m')}", DIAS_ABREV.get(dd.weekday(),"")
        except Exception:
            return data_str, ""

    # Conjunto de departamentos (ordenado) e estrutura de lookup
    deptos_ordem = {}
    for r in rows:
        dep = r["depto_nome"]
        if dep not in deptos_ordem:
            deptos_ordem[dep] = {"icone": r.get("depto_icone","👥"), "ordem": r.get("depto_ordem",999)}
    deptos_lista = sorted(deptos_ordem.keys(), key=lambda x: (deptos_ordem[x]["ordem"], x))

    # lookup[(depto, data, periodo)] = responsavel
    lookup = {}
    for r in rows:
        lookup[(r["depto_nome"], r["culto_data"], r["culto_periodo"])] = r.get("responsavel","")

    # Monta uma tabela-grade por período (Manhã / Noite)
    periodos_render = [per] if per else ["Manhã","Noite"]

    def montar_grade(periodo_alvo):
        # Datas únicas que têm esse período (ordenadas)
        datas_periodo = sorted({r["culto_data"] for r in rows if r["culto_periodo"]==periodo_alvo})
        if not datas_periodo:
            return ""
        # Cabeçalho de colunas: data + dia da semana
        ths = '<th class="dep-col">DEPARTAMENTO</th>'
        for dts in datas_periodo:
            dnum, dsem = label_data(dts)
            ths += f'<th><div class="dt">{dnum}</div><div class="ds">{dsem}</div></th>'
        # Linhas: um departamento por linha
        trs = ""
        for i, dep in enumerate(deptos_lista):
            cells = f'<td class="dep-col">{deptos_ordem[dep]["icone"]} {dep}</td>'
            for dts in datas_periodo:
                resp = lookup.get((dep, dts, periodo_alvo), "")
                if resp:
                    cells += f'<td class="resp">{resp}</td>'
                else:
                    cells += '<td class="vago">—</td>'
            bg = "#F8FAFF" if i % 2 else "#FFFFFF"
            trs += f'<tr style="background:{bg}">{cells}</tr>'
        per_cor = {"Manhã":"#F59E0B","Noite":"#6366F1"}.get(periodo_alvo,"#1B4FA8")
        return f"""<div class="grade-bloco">
          <div class="grade-titulo" style="background:{per_cor}">
            {"☀️" if periodo_alvo=="Manhã" else "🌙"} ESCALA — {titulo_mes.upper()} · {periodo_alvo.upper()}
          </div>
          <div class="grade-scroll">
            <table class="grade">
              <thead><tr>{ths}</tr></thead>
              <tbody>{trs}</tbody>
            </table>
          </div>
        </div>"""

    grade_html = ""
    for pr in periodos_render:
        grade_html += montar_grade(pr)
    if not grade_html:
        grade_html = '<p style="text-align:center;color:#94A3B8;padding:40px">Nenhuma escala cadastrada para este mês.</p>'

    html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Escala — {titulo_mes}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;background:#F1F5F9;color:#1E293B}}
.btn-print{{display:block;width:calc(100% - 32px);margin:16px auto 0;padding:14px;background:#0A2463;color:#fff;border:none;border-radius:12px;font-size:15px;font-weight:700;cursor:pointer}}
.hint{{text-align:center;font-size:11px;color:#94A3B8;margin:6px 16px 14px}}
.header{{background:linear-gradient(135deg,#0A2463,#1B4FA8);color:#fff;padding:22px 16px;margin-bottom:16px}}
.header h1{{font-size:21px;font-weight:900;letter-spacing:1px}}
.header .meta{{font-size:12px;opacity:.7;margin-top:4px}}
.grade-bloco{{background:#fff;border-radius:12px;margin:0 16px 18px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.grade-titulo{{color:#fff;font-size:13px;font-weight:800;padding:10px 14px;letter-spacing:.5px}}
.grade-scroll{{overflow-x:auto}}
table.grade{{width:100%;border-collapse:collapse;font-size:11px;min-width:520px}}
table.grade th{{background:#0F2747;color:#fff;padding:7px 9px;text-align:center;font-weight:700;border:1px solid #1B3357;white-space:nowrap}}
table.grade th.dep-col{{text-align:left;background:#0A1F3D;min-width:130px;position:sticky;left:0;z-index:2}}
table.grade th .dt{{font-size:12px}}
table.grade th .ds{{font-size:9px;opacity:.7;font-weight:600}}
table.grade td{{padding:6px 9px;text-align:center;border:1px solid #E2E8F0}}
table.grade td.dep-col{{text-align:left;font-weight:700;color:#0A2463;background:#EEF2F9;position:sticky;left:0;z-index:1;white-space:nowrap}}
table.grade td.resp{{font-weight:600;color:#0F2747}}
table.grade td.vago{{color:#CBD5E1}}
.footer{{text-align:center;font-size:10px;color:#94A3B8;padding:16px}}
@media print{{.btn-print,.hint{{display:none}}body{{background:#fff}}.grade-scroll{{overflow:visible}}table.grade{{min-width:0;font-size:9px}}}}
</style></head><body>
<button class="btn-print" onclick="window.print()">💾 Salvar como PDF</button>
<p class="hint">Toque nos 3 pontos → Imprimir → Salvar como PDF</p>
<div class="header">
  <h1>ESCALA — {titulo_mes.upper()}</h1>
  <div class="meta">Igreja ABA · {len(deptos_lista)} departamentos · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
</div>
{grade_html}
<div class="footer">Igreja ABA — Um Lar Para Pertencer</div>
</body></html>"""
    resp = make_response(html)
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    return resp

@app.route("/api/escala", methods=["POST"])
@login_required
def salvar_escala_item():
    d = request.get_json(force=True) or {}
    did  = d.get("departamento_id")
    # Apenas lider_depto tem restrição — admin/lider salvam qualquer depto
    cargo = session.get("usuario_cargo","")
    depto_filtro = get_depto_usuario() if cargo == "lider_depto" else None
    if depto_filtro and str(did) != str(depto_filtro):
        return jsonify({"erro":"Sem permissão para editar este departamento"}),403
    data = d.get("culto_data","").strip()
    per  = d.get("culto_periodo","").strip()
    resp = d.get("responsavel","").strip()
    if not did or not data or not per:
        return jsonify({"erro":"departamento_id, culto_data e culto_periodo são obrigatórios"}),400
    with get_db() as conn:
        # Upsert: atualiza se já existe, insere se não
        existe = conn.execute(
            qmark("SELECT id FROM escala_itens WHERE departamento_id=? AND culto_data=? AND culto_periodo=?"),
            (did, data, per)
        ).fetchone()
        if existe:
            conn.execute(
                qmark("UPDATE escala_itens SET responsavel=?,observacao=?,atualizado_em=? WHERE id=?"),
                (resp, d.get("observacao",""), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), existe["id"])
            )
        else:
            conn.execute(
                qmark("INSERT INTO escala_itens(departamento_id,culto_data,culto_periodo,responsavel,observacao) VALUES(?,?,?,?,?)"),
                (did, data, per, resp, d.get("observacao",""))
            )
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/escala/lote", methods=["POST"])
@login_required
def salvar_escala_lote():
    """Salva múltiplos itens de uma vez"""
    itens = request.get_json(force=True) or []
    if not isinstance(itens, list): return jsonify({"erro":"Lista esperada"}),400
    cargo = session.get("usuario_cargo","")
    depto_filtro = get_depto_usuario() if cargo == "lider_depto" else None
    salvos = 0
    with get_db() as conn:
        for d in itens:
            if depto_filtro and str(d.get("departamento_id")) != str(depto_filtro):
                continue
            did  = d.get("departamento_id")
            data = d.get("culto_data","").strip()
            per  = d.get("culto_periodo","").strip()
            resp = d.get("responsavel","").strip()
            if not did or not data or not per: continue
            existe = conn.execute(
                qmark("SELECT id FROM escala_itens WHERE departamento_id=? AND culto_data=? AND culto_periodo=?"),
                (did,data,per)
            ).fetchone()
            if existe:
                conn.execute(
                    qmark("UPDATE escala_itens SET responsavel=?,observacao=?,atualizado_em=? WHERE id=?"),
                    (resp, d.get("observacao",""), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), existe["id"])
                )
            else:
                conn.execute(
                    qmark("INSERT INTO escala_itens(departamento_id,culto_data,culto_periodo,responsavel,observacao) VALUES(?,?,?,?,?)"),
                    (did,data,per,resp,d.get("observacao",""))
                )
            salvos += 1
        conn.commit()
    return jsonify({"ok":True,"salvos":salvos})

@app.route("/api/escala/datas", methods=["GET"])
@login_required
def datas_culto_escala():
    """Retorna datas de cultos de um mês para montar a escala"""
    mes = request.args.get("mes", datetime.now().strftime("%Y-%m"))
    from calendar import monthrange
    from datetime import date as dt
    ano, m = int(mes.split("-")[0]), int(mes.split("-")[1])
    _, dias = monthrange(ano, m)
    datas = []
    for dia in range(1, dias+1):
        d = dt(ano, m, dia)
        wd = d.weekday()  # 4=sexta, 5=sábado, 6=domingo
        # Sexta e Sábado: só Noite | Domingo: Manhã e Noite
        if wd == 6:
            periodos = ["Manhã", "Noite"]
        elif wd in (4, 5):
            periodos = ["Noite"]
        else:
            periodos = []
        for per in periodos:
            datas.append({
                "data": d.isoformat(),
                "data_br": d.strftime("%d/%m"),
                "dia_semana": DIAS[wd],
                "periodo": per
            })
    return jsonify(datas)

# ═══════════════════════════════════════════════════════════════
# RELATÓRIOS DE GC
# ═══════════════════════════════════════════════════════════════

@app.route("/relatorio-gc")
def pagina_relatorio_gc():
    """Página de login exclusiva para líderes de GC"""
    return render_template("gc_relatorio_form.html")

@app.route("/api/relatorios_gc", methods=["POST"])
def criar_relatorio_gc():
    """Cria relatório semanal de GC — acesso por PIN do líder"""
    d = request.get_json(force=True) or {}
    gc_nome    = d.get("gc_nome","").strip()
    lider_nome = d.get("lider_nome","").strip()
    dia        = d.get("dia","").strip()
    if not gc_nome or not lider_nome or not dia:
        return jsonify({"erro":"GC, líder e dia são obrigatórios"}),400

    # Validação numérica de oferta e quilos (sem texto livre)
    def num(v):
        try:
            x = float(str(v).replace(",", ".").replace("R$","").strip() or 0)
            return max(0, x)
        except Exception:
            return 0
    valor_oferta = num(d.get("valor_oferta", 0))
    quilos       = num(d.get("quilos_arrecadados", 0))

    # Busca gc_id pelo nome
    with get_db() as conn:
        gc = conn.execute(qmark("SELECT id FROM grupos_crescimento WHERE nome=?"), (gc_nome,)).fetchone()
        gc_id = gc["id"] if gc else None

        conn.execute(
            qmark("""INSERT INTO relatorios_gc
                (gc_id,gc_nome,lider_nome,anfitriao,dia,membros_presentes,
                 visitantes,lider_treinamento,nome_lider_trein,observacoes,
                 valor_oferta,quilos_arrecadados)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)"""),
            (gc_id, gc_nome, lider_nome,
             d.get("anfitriao",""),
             dia,
             int(d.get("membros_presentes",0)),
             int(d.get("visitantes",0)),
             1 if d.get("lider_treinamento") else 0,
             d.get("nome_lider_trein",""),
             d.get("observacoes",""),
             valor_oferta, quilos)
        )
        conn.commit()
    return jsonify({"ok":True, "msg":"Relatório enviado com sucesso!"})

@app.route("/api/relatorios_gc", methods=["GET"])
@role_required("admin","lider")
def listar_relatorios_gc():
    """Lista relatórios de GC — apenas admin e lider do sistema"""
    ini = request.args.get("data_ini","")
    fim = request.args.get("data_fim","")
    gc  = request.args.get("gc_nome","")
    sql = "SELECT * FROM relatorios_gc WHERE 1=1"; p=[]
    if ini: sql+=" AND dia>=?"; p.append(ini)
    if fim: sql+=" AND dia<=?"; p.append(fim)
    if gc:  sql+=" AND gc_nome=?"; p.append(gc)
    sql+=" ORDER BY dia DESC, criado_em DESC"
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]
    for r in rows:
        r["dia_br"] = br(r["dia"]) if r.get("dia") else ""
    return jsonify(rows)

@app.route("/api/relatorios_gc/dashboard", methods=["GET"])
@role_required("admin","lider")
def dashboard_gc():
    """Dashboard consolidado dos relatórios de GC"""
    with get_db() as conn:
        # Total por GC
        por_gc = [dict(r) for r in conn.execute("""
            SELECT gc_nome,
                   COUNT(*) as total_reunioes,
                   COALESCE(SUM(membros_presentes),0) as total_membros,
                   COALESCE(SUM(visitantes),0) as total_visitantes,
                   COALESCE(ROUND(CAST(AVG(membros_presentes) AS NUMERIC),1),0) as media_membros,
                   COALESCE(ROUND(CAST(AVG(visitantes) AS NUMERIC),1),0) as media_visitantes,
                   MAX(dia) as ultima_reuniao
            FROM relatorios_gc
            GROUP BY gc_nome ORDER BY total_membros DESC
        """ if USE_PG else """
            SELECT gc_nome,
                   COUNT(*) as total_reunioes,
                   COALESCE(SUM(membros_presentes),0) as total_membros,
                   COALESCE(SUM(visitantes),0) as total_visitantes,
                   ROUND(AVG(membros_presentes),1) as media_membros,
                   ROUND(AVG(visitantes),1) as media_visitantes,
                   MAX(dia) as ultima_reuniao
            FROM relatorios_gc
            GROUP BY gc_nome ORDER BY total_membros DESC
        """).fetchall()]

        # Totais gerais
        totais = dict(conn.execute("""
            SELECT COUNT(*) as total_relatorios,
                   COALESCE(SUM(membros_presentes),0) as total_membros,
                   COALESCE(SUM(visitantes),0) as total_visitantes,
                   COALESCE(SUM(lider_treinamento),0) as total_lideres_trein
            FROM relatorios_gc
        """).fetchone() or {})

        # Últimos 10 relatórios
        ultimos = [dict(r) for r in conn.execute(
            "SELECT * FROM relatorios_gc ORDER BY dia DESC, criado_em DESC LIMIT 10"
        ).fetchall()]

    def safe(v):
        if v is None: return 0
        try: return float(v) if "." in str(v) else int(v)
        except: return 0

    totais = {k: safe(v) if k!="gc_nome" else v for k,v in totais.items()}
    por_gc_clean = [{k: safe(v) if k not in ("gc_nome","ultima_reuniao") else v for k,v in g.items()} for g in por_gc]

    for r in ultimos:
        r["dia_br"] = br(r["dia"]) if r.get("dia") else ""

    return jsonify({
        "totais":   totais,
        "por_gc":   por_gc_clean,
        "ultimos":  ultimos
    })

@app.route("/api/relatorios_gc/<int:rid>", methods=["DELETE"])
@role_required("admin","lider")
def deletar_relatorio_gc(rid):
    with get_db() as conn:
        conn.execute(qmark("DELETE FROM relatorios_gc WHERE id=?"), (rid,))
        conn.commit()
    return jsonify({"ok":True})

# ═══════════════════════════════════════════════════════════════
# RELATÓRIO DE FREQUÊNCIA DOS GCs
# ═══════════════════════════════════════════════════════════════
@app.route("/api/relatorios_gc/frequencia", methods=["GET"])
@role_required("admin","lider")
def frequencia_gc():
    """Relatório de frequência por GC com média, filtros e exportação"""
    gc_nome = request.args.get("gc_nome","")
    ini     = request.args.get("data_ini","")
    fim     = request.args.get("data_fim","")
    sql = "SELECT * FROM relatorios_gc WHERE 1=1"; p=[]
    if gc_nome: sql+=" AND gc_nome=?"; p.append(gc_nome)
    if ini: sql+=" AND dia>=?"; p.append(ini)
    if fim: sql+=" AND dia<=?"; p.append(fim)
    sql+=" ORDER BY gc_nome,dia"
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]
        gcs  = [dict(r) for r in conn.execute("SELECT nome FROM grupos_crescimento WHERE ativo=1 ORDER BY nome").fetchall()]

    # Calcula frequência por GC
    por_gc = {}
    for r in rows:
        nome = r["gc_nome"]
        if nome not in por_gc:
            por_gc[nome] = {"gc_nome":nome,"reunioes":0,"total_membros":0,"total_visitantes":0,"datas":[]}
        por_gc[nome]["reunioes"] += 1
        por_gc[nome]["total_membros"] += r.get("membros_presentes",0)
        por_gc[nome]["total_visitantes"] += r.get("visitantes",0)
        por_gc[nome]["datas"].append({
            "dia": r["dia"],
            "dia_br": br(r["dia"]),
            "membros": r.get("membros_presentes",0),
            "visitantes": r.get("visitantes",0),
            "lider": r.get("lider_nome",""),
            "anfitriao": r.get("anfitriao",""),
            "observacoes": r.get("observacoes","")
        })

    resultado = []
    for gc_nome_key, g in por_gc.items():
        n = max(g["reunioes"],1)
        resultado.append({
            **g,
            "media_membros": round(g["total_membros"]/n,1),
            "media_visitantes": round(g["total_visitantes"]/n,1)
        })
    resultado.sort(key=lambda x: x["total_membros"], reverse=True)

    def safe(v):
        if v is None: return 0
        try: return float(v) if "." in str(v) else int(v)
        except: return 0

    totais = {
        "total_relatorios": len(rows),
        "total_gcs": len(por_gc),
        "media_geral_membros": round(sum(g["total_membros"] for g in por_gc.values())/max(len(rows),1),1),
        "media_geral_visitantes": round(sum(g["total_visitantes"] for g in por_gc.values())/max(len(rows),1),1)
    }
    return jsonify({"ok":True,"totais":totais,"por_gc":resultado,"gcs_lista":[g["nome"] for g in gcs]})

@app.route("/api/relatorios_gc/frequencia/pdf", methods=["GET"])
@role_required("admin","lider")
def frequencia_gc_pdf():
    """PDF de frequência dos GCs"""
    gc_nome = request.args.get("gc_nome","")
    ini     = request.args.get("data_ini","")
    fim     = request.args.get("data_fim","")

    sql = "SELECT * FROM relatorios_gc WHERE 1=1"; p=[]
    if gc_nome: sql+=" AND gc_nome=?"; p.append(gc_nome)
    if ini: sql+=" AND dia>=?"; p.append(ini)
    if fim: sql+=" AND dia<=?"; p.append(fim)
    sql+=" ORDER BY gc_nome,dia"
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]

    por_gc = {}
    for r in rows:
        nome = r["gc_nome"]
        if nome not in por_gc: por_gc[nome]={"reunioes":0,"total_membros":0,"total_visitantes":0,"datas":[]}
        por_gc[nome]["reunioes"]+=1
        por_gc[nome]["total_membros"]+=r.get("membros_presentes",0)
        por_gc[nome]["total_visitantes"]+=r.get("visitantes",0)
        por_gc[nome]["datas"].append(r)

    cards_html=""
    for nome,g in sorted(por_gc.items(),key=lambda x:x[1]["total_membros"],reverse=True):
        n=max(g["reunioes"],1)
        media=round(g["total_membros"]/n,1)
        media_v=round(g["total_visitantes"]/n,1)
        rows_html="".join(f"<tr><td>{br(d['dia'])}</td><td style='text-align:center'>{d.get('membros_presentes',0)}</td><td style='text-align:center'>{d.get('visitantes',0)}</td><td>{d.get('anfitriao','')}</td></tr>" for d in g["datas"])
        cards_html+=f"""<div class="gc-card">
          <div class="gc-header"><span class="gc-nome">{nome}</span>
            <div class="gc-badges">
              <span class="badge blue">{g['reunioes']} reuniões</span>
              <span class="badge green">Média: {media} membros</span>
            </div>
          </div>
          <table class="mini-table"><thead><tr><th>Data</th><th>Membros</th><th>Visitantes</th><th>Anfitrião</th></tr></thead>
          <tbody>{rows_html}</tbody></table>
          <div class="gc-totais">Total membros: <strong>{g['total_membros']}</strong> | Total visitantes: <strong>{g['total_visitantes']}</strong> | Média visit.: <strong>{media_v}</strong></div>
        </div>"""

    total_membros=sum(g["total_membros"] for g in por_gc.values())
    total_vis=sum(g["total_visitantes"] for g in por_gc.values())
    filtro=""
    if gc_nome: filtro+=f" | GC: {gc_nome}"
    if ini or fim: filtro+=f" | {br(ini) if ini else '...'} — {br(fim) if fim else '...'}"

    html=f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Frequência GCs — Igreja ABA</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;background:#F1F5F9;color:#1E293B;padding:0}}
.btn-print{{display:block;width:calc(100% - 32px);margin:16px auto 0;padding:14px;background:#0A2463;color:#fff;border:none;border-radius:12px;font-size:15px;font-weight:700;cursor:pointer}}
.hint{{text-align:center;font-size:11px;color:#94A3B8;margin:6px 16px 14px}}
.header{{background:linear-gradient(135deg,#0A2463,#1B4FA8);color:#fff;padding:20px 16px;margin:0 0 16px}}
.header h1{{font-size:20px;font-weight:900;letter-spacing:2px}}
.header .meta{{font-size:11px;opacity:.6;margin-top:4px}}
.stats{{display:grid;grid-template-columns:1fr 1fr;gap:10px;padding:0 16px 16px}}
.stat{{background:#fff;border-radius:12px;padding:12px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.stat.blue{{background:linear-gradient(135deg,#0A2463,#1B4FA8)}}
.stat .sv{{font-size:26px;font-weight:800;color:#0A2463}}
.stat.blue .sv{{color:#fff}}
.stat .sl{{font-size:10px;color:#64748B;text-transform:uppercase;margin-top:3px}}
.stat.blue .sl{{color:rgba(255,255,255,.7)}}
.gc-card{{background:#fff;border-radius:12px;margin:0 16px 12px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.gc-header{{display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:6px;margin-bottom:10px}}
.gc-nome{{font-size:14px;font-weight:800;color:#0A2463}}
.gc-badges{{display:flex;gap:5px;flex-wrap:wrap}}
.badge{{padding:2px 8px;border-radius:20px;font-size:10px;font-weight:700;color:#fff}}
.badge.blue{{background:#1B4FA8}}.badge.green{{background:#059669}}
.mini-table{{width:100%;border-collapse:collapse;font-size:11px;margin-bottom:8px}}
.mini-table th{{background:#F1F5F9;padding:5px 8px;text-align:left;font-weight:600;color:#4A6080}}
.mini-table td{{padding:5px 8px;border-bottom:1px solid #F1F5F9}}
.gc-totais{{font-size:11px;color:#4A6080;background:#F8FAFF;border-radius:8px;padding:6px 10px}}
.footer{{text-align:center;font-size:10px;color:#94A3B8;padding:16px}}
@media print{{.btn-print,.hint{{display:none}}body{{background:#fff}}}}
</style></head><body>
<button class="btn-print" onclick="window.print()">💾 Salvar como PDF</button>
<p class="hint">Toque nos 3 pontos → Imprimir</p>
<div class="header">
  <h1>FREQUÊNCIA DOS GCs</h1>
  <div class="meta">Igreja ABA · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}{filtro}</div>
</div>
<div class="stats">
  <div class="stat blue"><div class="sv">{len(rows)}</div><div class="sl">Relatórios</div></div>
  <div class="stat blue"><div class="sv">{len(por_gc)}</div><div class="sl">GCs Ativos</div></div>
  <div class="stat"><div class="sv" style="color:#059669">{total_membros}</div><div class="sl">Total Membros</div></div>
  <div class="stat"><div class="sv" style="color:#7C3AED">{total_vis}</div><div class="sl">Total Visitantes</div></div>
</div>
{cards_html}
<div class="footer">Igreja ABA — Um Lar Para Pertencer</div>
</body></html>"""
    resp=make_response(html)
    resp.headers["Content-Type"]="text/html; charset=utf-8"
    return resp

@app.route("/api/relatorios_gc/frequencia/excel", methods=["GET"])
@role_required("admin","lider")
def frequencia_gc_excel():
    """Excel completo de frequência dos GCs"""
    gc_filtro = request.args.get("gc_nome","")
    ini       = request.args.get("data_ini","")
    fim       = request.args.get("data_fim","")
    sql = "SELECT * FROM relatorios_gc WHERE 1=1"; p=[]
    if gc_filtro: sql+=" AND gc_nome=?"; p.append(gc_filtro)
    if ini: sql+=" AND dia>=?"; p.append(ini)
    if fim: sql+=" AND dia<=?"; p.append(fim)
    sql+=" ORDER BY gc_nome,dia"
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(qmark(sql),p).fetchall()]

    wb = openpyxl.Workbook()
    azul   = PatternFill("solid",fgColor="0A2463")
    azul2  = PatternFill("solid",fgColor="1B4FA8")
    cinza  = PatternFill("solid",fgColor="EBF5FF")
    verde  = PatternFill("solid",fgColor="D1FAE5")
    amarelo= PatternFill("solid",fgColor="FEF3C7")
    branco = PatternFill("solid",fgColor="FFFFFF")
    tw = Font(color="FFFFFF",bold=True,name="Calibri",size=11)
    tn = Font(name="Calibri",size=10)
    tb = Font(name="Calibri",size=10,bold=True)
    centro=Alignment(horizontal="center",vertical="center")
    esq=Alignment(horizontal="left",vertical="center",wrap_text=True)
    borda=Border(left=Side(style="thin",color="D1D5DB"),right=Side(style="thin",color="D1D5DB"),
                 top=Side(style="thin",color="D1D5DB"),bottom=Side(style="thin",color="D1D5DB"))

    def cab(ws,row,cols,fill=None):
        fill=fill or azul
        for i,t in enumerate(cols,1):
            c=ws.cell(row=row,column=i,value=t)
            c.fill=fill;c.font=tw;c.alignment=centro;c.border=borda
    def cel(ws,row,col,val,bold=False,fill=None,align=None):
        c=ws.cell(row=row,column=col,value=val)
        c.fill=fill or branco;c.font=tb if bold else tn
        c.alignment=align or centro;c.border=borda;return c

    # Aba Resumo por GC
    ws1=wb.active; ws1.title="Resumo por GC"
    ws1.merge_cells("A1:F1")
    c=ws1["A1"];c.value="FREQUÊNCIA DOS GCS — IGREJA ABA";c.fill=azul
    c.font=Font(color="FFFFFF",bold=True,name="Calibri",size=14);c.alignment=centro
    cab(ws1,2,["GC","Reuniões","Total Membros","Total Visitantes","Média Membros","Média Visitantes"],azul)
    larguras={"A":28,"B":12,"C":16,"D":16,"E":16,"F":16}
    for col,w in larguras.items(): ws1.column_dimensions[col].width=w

    por_gc={}
    for r in rows:
        n=r["gc_nome"]
        if n not in por_gc: por_gc[n]={"reunioes":0,"membros":0,"vis":0}
        por_gc[n]["reunioes"]+=1
        por_gc[n]["membros"]+=r.get("membros_presentes",0)
        por_gc[n]["vis"]+=r.get("visitantes",0)

    for i,(nome,g) in enumerate(sorted(por_gc.items(),key=lambda x:x[1]["membros"],reverse=True),3):
        n=max(g["reunioes"],1)
        fill=cinza if i%2==0 else branco
        cel(ws1,i,1,nome,bold=True,fill=fill,align=esq)
        cel(ws1,i,2,g["reunioes"],fill=fill)
        cel(ws1,i,3,g["membros"],fill=fill)
        cel(ws1,i,4,g["vis"],fill=fill)
        cel(ws1,i,5,round(g["membros"]/n,1),fill=fill)
        cel(ws1,i,6,round(g["vis"]/n,1),fill=fill)

    # Linha totais
    tot=len(por_gc)+3
    cel(ws1,tot,1,"MÉDIA GERAL",bold=True,fill=amarelo,align=esq)
    cel(ws1,tot,5,round(sum(g["membros"] for g in por_gc.values())/max(len(rows),1),1),bold=True,fill=amarelo)
    cel(ws1,tot,6,round(sum(g["vis"] for g in por_gc.values())/max(len(rows),1),1),bold=True,fill=amarelo)

    # Aba detalhe
    ws2=wb.create_sheet("Todos Relatórios")
    cols2=["Data","GC","Líder","Anfitrião","Membros","Visitantes","Líder Trein.","Nome Líder Trein.","Observações"]
    larg2=[12,24,20,20,10,10,12,22,35]
    for i,(c,l) in enumerate(zip(cols2,larg2),1): ws2.column_dimensions[get_column_letter(i)].width=l
    ws2.row_dimensions[1].height=28
    cab(ws2,1,cols2)
    for i,r in enumerate(rows,2):
        fill=cinza if i%2==0 else branco
        vals=[br(r.get("dia","")),r.get("gc_nome",""),r.get("lider_nome",""),
              r.get("anfitriao",""),r.get("membros_presentes",0),r.get("visitantes",0),
              "Sim" if r.get("lider_treinamento") else "Não",r.get("nome_lider_trein",""),r.get("observacoes","")]
        for j,v in enumerate(vals,1):
            cel(ws2,i,j,v,fill=fill,align=esq if j in (2,3,4,8,9) else centro)

    tot2=len(rows)+2
    cel(ws2,tot2,1,"TOTAIS",bold=True,fill=amarelo)
    ws2.merge_cells(f"A{tot2}:D{tot2}")
    cel(ws2,tot2,5,sum(r.get("membros_presentes",0) for r in rows),bold=True,fill=amarelo)
    cel(ws2,tot2,6,sum(r.get("visitantes",0) for r in rows),bold=True,fill=amarelo)
    media2=tot2+1
    cel(ws2,media2,1,"MÉDIAS",bold=True,fill=verde)
    ws2.merge_cells(f"A{media2}:D{media2}")
    n2=max(len(rows),1)
    cel(ws2,media2,5,round(sum(r.get("membros_presentes",0) for r in rows)/n2,1),bold=True,fill=verde)
    cel(ws2,media2,6,round(sum(r.get("visitantes",0) for r in rows)/n2,1),bold=True,fill=verde)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,
                     download_name=f"frequencia_gc_{date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════════════════════
# INIT
# ═══════════════════════════════════════════════════════════════
init_db()

def limpar_escalas_nao_domingo():
    """Remove itens de escala em dias que não são de culto (mantém sexta, sábado e domingo)"""
    try:
        with get_db() as conn:
            rows = [dict(r) for r in conn.execute("SELECT id,culto_data FROM escala_itens").fetchall()]
            ids_remover = [r["id"] for r in rows if not _eh_domingo(r.get("culto_data",""))]
            for rid in ids_remover:
                conn.execute(qmark("DELETE FROM escala_itens WHERE id=?"), (rid,))
            if ids_remover:
                conn.commit()
                logger.info(f"Limpeza: {len(ids_remover)} itens de escala em dias inválidos removidos")
    except Exception as e:
        logger.warning(f"Limpeza escalas falhou: {e}")

limpar_escalas_nao_domingo()

if __name__=="__main__":
    print(f"  IGREJA ABA v5 | DB: {DB_PATH}")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT",5000)), debug=False)
